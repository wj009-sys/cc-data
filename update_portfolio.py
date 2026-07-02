#!/usr/bin/env python
"""
400万资产配置组合 - 每日自动更新脚本
用于 Windows 任务计划程序，在每个交易日 16:30 执行

功能：
1. 判断是否为交易日（含中国节假日）
2. 从 Tushare 获取所有持仓 ETF 最新收盘价
3. 更新「汇总」sheet：收盘价、浮盈/亏、权重、偏离、信号
4. 追加「日报」sheet：当日所有品种行情快照
5. 根据投资纪律检测并标注触发信号
"""
import os
import sys
import io
import json
import pickle
from datetime import datetime, date, timedelta
from pathlib import Path

# 修复 Windows GBK 编码问题
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# 绕过系统代理（Tushare API 直连）
os.environ['NO_PROXY'] = 'api.waditu.com,*.waditu.com,localhost,127.0.0.1'

# ==== 配置 ====
FILE_PATH = Path(r'C:\Users\65004\Desktop\小白\cc-data\400万资产配置组合2026.xlsx')
CACHE_PATH = Path(r'C:\Users\65004\Desktop\小白\cc-data\.portfolio_cache.pkl')

# 中国节假日（2026年，需年末更新下一年）
CN_HOLIDAYS_2026 = {
    date(2026, 1, 1), date(2026, 1, 2),                                    # 元旦
    date(2026, 2, 16), date(2026, 2, 17), date(2026, 2, 18),               # 春节(2.17除夕)
    date(2026, 2, 19), date(2026, 2, 20),
    date(2026, 4, 6),                                                        # 清明
    date(2026, 5, 1), date(2026, 5, 4), date(2026, 5, 5),                  # 劳动节
    date(2026, 6, 19),                                                       # 端午(暂)
    date(2026, 9, 25),                                                       # 中秋(暂)
    date(2026, 10, 1), date(2026, 10, 2), date(2026, 10, 5),               # 国庆
    date(2026, 10, 6), date(2026, 10, 7),
}
# 调休工作日（周六日补班）
CN_WORKDAYS_2026 = set()

# 持仓定义：代码 -> (名称, tushare_code, 高波动品种, 权益/固收/商品分类)
HOLDINGS = [
    {"code": 510300, "name": "沪深300ETF",        "ts": "510300.SH", "high_vol": False, "type": "equity"},
    {"code": 588050, "name": "科创50ETF",         "ts": "588050.SH", "high_vol": True,  "type": "equity"},
    {"code": 159915, "name": "创业板ETF",         "ts": "159915.SZ", "high_vol": True,  "type": "equity"},
    {"code": 512890, "name": "红利低波ETF",       "ts": "512890.SH", "high_vol": False, "type": "equity"},
    {"code": 511380, "name": "可转债ETF",         "ts": "511380.SH", "high_vol": False, "type": "equity"},
    {"code": 511260, "name": "10年国债ETF",       "ts": "511260.SH", "high_vol": False, "type": "bond"},
    {"code": 511010, "name": "5年国债ETF",        "ts": "511010.SH", "high_vol": False, "type": "bond"},
    {"code": 511360, "name": "短融ETF",           "ts": "511360.SH", "high_vol": False, "type": "bond"},
    {"code": 518880, "name": "黄金ETF",           "ts": "518880.SH", "high_vol": False, "type": "commodity"},
]

HOLDING_MAP = {h["code"]: h for h in HOLDINGS}
# 汇总 sheet 中每个代码对应的行号（第2行开始，与 HOLDINGS 顺序一致）
CODE_ROWS = [2, 3, 4, 5, 6, 7, 8, 9, 10]

# ==== 交易日判断 ====
def is_trading_day(d: date) -> bool:
    """判断是否为 A 股交易日"""
    if d in CN_HOLIDAYS_2026:
        return False
    if d.weekday() < 5:  # 周一至周五
        return True
    if d in CN_WORKDAYS_2026:  # 调休补班
        return True
    return False


def get_latest_trading_day() -> date:
    """获取最近一个交易日"""
    d = date.today()
    for _ in range(10):
        if is_trading_day(d):
            return d
        d = d - timedelta(days=1)
    return date.today()


# ==== Tushare 数据获取 ====
def fetch_prices(ts_codes: list, trade_date: str) -> dict:
    """批量获取 ETF 最新收盘价，返回 {ts_code: {close, pct_chg, pre_close}}"""
    import tushare as ts
    pro = ts.pro_api()
    result = {}

    for code in ts_codes:
        try:
            df = pro.fund_daily(ts_code=code, trade_date=trade_date)
            if not df.empty:
                row = df.iloc[0]
                result[code] = {
                    "close": float(row["close"]),
                    "pct_chg": float(row["pct_chg"]),
                    "pre_close": float(row["pre_close"]),
                }
            else:
                # 尝试扩大范围取最近数据
                start = (datetime.strptime(trade_date, "%Y%m%d") - timedelta(days=5)).strftime("%Y%m%d")
                df2 = pro.fund_daily(ts_code=code, start_date=start, end_date=trade_date)
                if not df2.empty:
                    row = df2.iloc[0]
                    result[code] = {
                        "close": float(row["close"]),
                        "pct_chg": float(row["pct_chg"]),
                        "pre_close": float(row["pre_close"]),
                    }
        except Exception as e:
            print(f"  [WARN] {code} 获取失败: {e}")

    return result


# ==== 信号检测 ====
def detect_signals(holdings_data: list, trade_date_str: str = None) -> dict:
    """
    根据投资纪要规则检测触发信号（V3优化版）。
    holdings_data: [{code, name, close, cost_price, quantity, weight, target_weight, pct_chg, type, high_vol}, ...]
    trade_date_str: YYYYMMDD 格式的交易日期（用于判断6月/12月末最后一个交易日）
    返回: {code: [signal_strings]}
    """
    signals = {}

    # 判断是否为半年度末检查日 (6月/12月最后一个交易日)
    # 注意：is_end_of_period 用于再平衡（仅12月触发）和潜在的半年度评估
    is_end_of_period = False
    if trade_date_str:
        dt = datetime.strptime(trade_date_str, "%Y%m%d")
        if dt.month in (6, 12):
            # 检查该月后续是否还有交易日
            check_day = dt.date()
            has_more_trading_days = False
            for d in range(1, 11):
                future = check_day + timedelta(days=d)
                if future.month != check_day.month:
                    break
                if is_trading_day(future):
                    has_more_trading_days = True
                    break
            if not has_more_trading_days:
                is_end_of_period = True

    for h in holdings_data:
        s = []
        code = h["code"]
        close = h["close"]
        cost = h.get("cost_price")
        qty = h.get("quantity", 0)
        weight = h.get("weight", 0)
        target = h.get("target_weight", 0)
        deviation = weight - target
        is_high_vol = h.get("high_vol", False)
        htype = h.get("type", "equity")

        if qty and qty > 0 and cost and cost > 0:
            pnl_pct = (close - cost) / cost

            # 一、V3统一止盈（+20%卖半，仅权益类，不含黄金/债）
            if htype == "equity" and pnl_pct >= 0.20:
                s.append(f"🟢V3止盈：浮盈{pnl_pct:.1%} >= +20%，卖半仓，资金转入511360短融ETF")

            # 二、V3统一止损（-15%观察20日，所有权益类统一，取消高波动专有止损）
            if htype == "equity":
                if pnl_pct <= -0.15:
                    s.append(f"🔴V3止损：浮亏{pnl_pct:.1%} <= -15%，次日清仓观察20日")

        # 三、V3再平衡：仅年度末（12月最后一个交易日）检查 ±20% 相对阀值
        # is_end_of_period 也覆盖6月但此处仅12月触发再平衡
        if is_end_of_period and dt.month == 12 and target > 0 and weight > 0:
            rel_dev = (weight - target) / target
            if abs(rel_dev) > 0.20:
                s.append(f"🔄V3年度再平衡：相对偏离{rel_dev:+.1%}，触发±20%阀值，调回目标权重")

        signals[code] = s

    return signals


def detect_portfolio_signals(total_pnl_pct: float, equity_weight: float) -> list:
    """四、V3熔断与黑天鹅应对"""
    s = []
    if total_pnl_pct <= -0.15:
        s.append("⚫黑天鹅：组合回撤>=15%，权益压至40%以下，保留20%现金+40%债券")
    elif total_pnl_pct <= -0.12:
        s.append("🔴组合熔断：回撤>=12%，权益压至60%以下，资金转债券/现金")
    elif total_pnl_pct <= -0.10:
        s.append("🟠组合预警：回撤>=10%，暂停新建仓+追涨，允许止盈/止损/再平衡")
    elif total_pnl_pct <= -0.05:
        s.append("🟡组合关注：回撤>=5%，维持现持仓观望")
    return s


# ==== 日报缺失交易日回补 ====
def normalize_date_str(val) -> str:
    """
    将日报 sheet 中各种日期格式统一转为 YYYYMMDD 字符串。
    支持: date/datetime 对象、8位数字 "20260626"、带横线 "2026-06-26" 等
    """
    if isinstance(val, (date, datetime)):
        return val.strftime("%Y%m%d")
    s = str(val).strip()
    # 8位数字
    if len(s) == 8 and s.isdigit():
        return s
    # 2026-06-26 或 2026-06-26 00:00:00
    digits = [c for c in s if c.isdigit()]
    if len(digits) >= 8:
        return "".join(digits[:8])
    return s  # fallback


def backfill_missing_trading_days(wb, ws_summary, ws_daily, latest_trade_date_str):
    """
    检查日报中缺失的交易日，自动回补。
    防止因脚本漏跑导致走势图断档。
    """
    # 收集日报已有日期（统一用 YYYYMMDD 格式）
    existing_dates = set()
    for row in range(2, ws_daily.max_row + 1):
        dval = ws_daily.cell(row=row, column=1).value
        if dval:
            dnorm = normalize_date_str(dval)
            if len(dnorm) == 8 and dnorm.isdigit():
                existing_dates.add(dnorm)

    if not existing_dates:
        return

    latest_dt = datetime.strptime(latest_trade_date_str, "%Y%m%d").date()
    earliest_str = min(existing_dates)
    earliest_dt = datetime.strptime(earliest_str, "%Y%m%d").date()

    # 找出缺失的交易日
    missing_dates = []
    d = earliest_dt + timedelta(days=1)
    while d <= latest_dt:
        d_str = d.strftime("%Y%m%d")
        if is_trading_day(d) and d_str not in existing_dates:
            missing_dates.append(d_str)
        d = d + timedelta(days=1)

    if not missing_dates:
        return

    # 安全限制：仅回补最近 10 个交易日内的缺失数据
    # 超过此范围的回补使用当前持仓数据（而非历史数据），会造成 P&L 失真
    ten_trading_days_ago = latest_dt
    count = 0
    while count < 10 and ten_trading_days_ago > earliest_dt:
        ten_trading_days_ago -= timedelta(days=1)
        if is_trading_day(ten_trading_days_ago):
            count += 1
    recent_missing = [d for d in missing_dates if d >= ten_trading_days_ago.strftime("%Y%m%d")]
    skipped = [d for d in missing_dates if d < ten_trading_days_ago.strftime("%Y%m%d")]
    if skipped:
        print(f"  [WARN] 跳过 {len(skipped)} 个超过10个交易日的缺失日期（历史数据回补会导致持仓数据失真）: {skipped}")
    if not recent_missing:
        print(f"  [INFO] 无可回补的近期缺失日期")
        return

    print(f"  [INFO] 检测到 {len(recent_missing)} 个缺失交易日（含 {len(skipped)} 个已跳过）: {recent_missing}")

    ts_codes = [h["ts"] for h in HOLDINGS]
    for d_str in recent_missing:
        print(f"    回补 {d_str}...")
        prices = fetch_prices(ts_codes, d_str)
        if not prices:
            print(f"      [WARN] {d_str} 无行情数据，跳过")
            continue

        # 汇总当日浮动盈亏
        total_mv = 0
        total_cost = 0
        rows_data = []
        for i, h in enumerate(HOLDINGS):
            row = CODE_ROWS[i]
            ts_code = h["ts"]
            cost_price = ws_summary.cell(row=row, column=8).value
            quantity = ws_summary.cell(row=row, column=9).value or 0
            target_weight = ws_summary.cell(row=row, column=5).value or 0

            if ts_code in prices:
                close = prices[ts_code]["close"]
                pct_chg = prices[ts_code]["pct_chg"]
            else:
                close = ws_summary.cell(row=row, column=11).value or 0
                pct_chg = 0

            if quantity and quantity > 0 and cost_price and cost_price > 0:
                mv = close * quantity
                cost_v = cost_price * quantity
                total_mv += mv
                total_cost += cost_v
                pnl = round((close - cost_price) * quantity, 2)
                pnl_pct = round((close - cost_price) / cost_price, 4)
            else:
                mv = 0
                cost_v = 0
                pnl = 0
                pnl_pct = 0

            rows_data.append({
                "code": h["code"], "name": h["name"], "close": close,
                "cost_price": cost_price, "quantity": quantity,
                "pct_chg": pct_chg, "pnl": pnl, "pnl_pct": pnl_pct,
                "market_value": mv, "cost_value": cost_v,
                "target_weight": target_weight,
            })

        # 计算权重
        for rd in rows_data:
            if total_mv > 0 and rd["quantity"] and rd["quantity"] > 0:
                rd["weight"] = round(rd["market_value"] / total_mv, 4)
            else:
                rd["weight"] = 0
            rd["deviation"] = round(rd["weight"] - rd["target_weight"], 4)

        # 删除旧数据（如有）并追加
        existing_rows = []
        for r in range(2, ws_daily.max_row + 1):
            dval_norm = normalize_date_str(ws_daily.cell(row=r, column=1).value)
            if dval_norm == d_str:
                existing_rows.append(r)
        for r in reversed(existing_rows):
            ws_daily.delete_rows(r)

        next_row = ws_daily.max_row + 1
        for rd in rows_data:
            ws_daily.cell(row=next_row, column=1).value = d_str
            ws_daily.cell(row=next_row, column=2).value = rd["name"]
            ws_daily.cell(row=next_row, column=3).value = str(rd["code"])
            ws_daily.cell(row=next_row, column=4).value = rd["quantity"] if rd["quantity"] else None
            ws_daily.cell(row=next_row, column=5).value = rd["cost_price"] if rd["cost_price"] else None
            ws_daily.cell(row=next_row, column=6).value = rd["cost_value"] if rd["cost_value"] else None
            ws_daily.cell(row=next_row, column=7).value = rd["close"]
            ws_daily.cell(row=next_row, column=8).value = round(rd["pct_chg"] / 100, 4) if rd["pct_chg"] else None
            ws_daily.cell(row=next_row, column=9).value = rd["market_value"]
            ws_daily.cell(row=next_row, column=10).value = rd["pnl"]
            ws_daily.cell(row=next_row, column=11).value = rd["pnl_pct"]
            ws_daily.cell(row=next_row, column=12).value = rd["target_weight"]
            ws_daily.cell(row=next_row, column=13).value = rd["weight"]
            ws_daily.cell(row=next_row, column=14).value = rd["deviation"]
            ws_daily.cell(row=next_row, column=15).value = "—"
            next_row += 1

        print(f"      已追加 {len(rows_data)} 个品种 ({d_str})")


# ==== Excel 更新 ====
def update_spreadsheet(prices: dict, trade_date_str: str):
    """主更新逻辑"""
    from openpyxl import load_workbook

    wb = load_workbook(FILE_PATH)
    ws_summary = wb["汇总"]
    ws_daily = wb["日报"]

    # --- 1. 更新汇总 sheet ---
    total_market_value = 0
    total_cost_value = 0
    holdings_data = []

    for i, h in enumerate(HOLDINGS):
        row = CODE_ROWS[i]
        ts_code = h["ts"]
        code = h["code"]
        cost_price = ws_summary.cell(row=row, column=8).value  # H: 成本价
        quantity = ws_summary.cell(row=row, column=9).value or 0  # I: 数量

        if ts_code in prices:
            close = prices[ts_code]["close"]
            pct_chg = prices[ts_code]["pct_chg"]
            # 更新 K: 今日收盘价
            ws_summary.cell(row=row, column=11).value = close

            if quantity and quantity > 0 and cost_price and cost_price > 0:
                cost_value = cost_price * quantity
                market_value = close * quantity
                total_market_value += market_value
                total_cost_value += cost_value

                holdings_data.append({
                    "code": code, "name": h["name"], "close": close,
                    "cost_price": cost_price, "quantity": quantity,
                    "target_weight": ws_summary.cell(row=row, column=5).value or 0,
                    "type": h["type"], "high_vol": h["high_vol"],
                    "pct_chg": pct_chg,
                })
            else:
                holdings_data.append({
                    "code": code, "name": h["name"], "close": close,
                    "cost_price": cost_price, "quantity": 0,
                    "target_weight": ws_summary.cell(row=row, column=5).value or 0,
                    "type": h["type"], "high_vol": h["high_vol"],
                    "pct_chg": pct_chg,
                })
        else:
            # 价格未获取，保留原值
            close = ws_summary.cell(row=row, column=11).value or 0
            cost_price = ws_summary.cell(row=row, column=8).value
            quantity = ws_summary.cell(row=row, column=9).value or 0
            if quantity and quantity > 0 and cost_price and cost_price > 0:
                market_value = close * quantity
                cost_value = cost_price * quantity
                total_market_value += market_value
                total_cost_value += cost_value

            holdings_data.append({
                "code": code, "name": h["name"], "close": close,
                "cost_price": cost_price, "quantity": quantity,
                "target_weight": ws_summary.cell(row=row, column=5).value or 0,
                "type": h["type"], "high_vol": h["high_vol"],
                "pct_chg": 0,
            })

    # 更新持仓权重和偏离
    for i, h in enumerate(HOLDINGS):
        row = CODE_ROWS[i]
        cost_price = ws_summary.cell(row=row, column=8).value
        quantity = ws_summary.cell(row=row, column=9).value or 0
        close = ws_summary.cell(row=row, column=11).value or 0
        target_weight = ws_summary.cell(row=row, column=5).value or 0

        if total_market_value > 0 and quantity and quantity > 0:
            actual_weight = round((close * quantity) / total_market_value, 4)
        else:
            actual_weight = 0

        ws_summary.cell(row=row, column=6).value = actual_weight  # F: 当前权重
        ws_summary.cell(row=row, column=7).value = round(actual_weight - target_weight, 4)  # G: 偏离

        # 更新 P&L 和 P&L%（统一处理，含零持仓）
        if quantity and quantity > 0 and cost_price and cost_price > 0:
            ws_summary.cell(row=row, column=12).value = round((close - cost_price) * quantity, 2)
            ws_summary.cell(row=row, column=13).value = round((close - cost_price) / cost_price, 4)
        else:
            ws_summary.cell(row=row, column=12).value = 0
            ws_summary.cell(row=row, column=13).value = 0

    # 更新表头日期
    today_str = datetime.now().strftime("%Y-%m-%d")
    trade_day_str = datetime.strptime(trade_date_str, "%Y%m%d").strftime("%Y-%m-%d")
    ws_summary.cell(row=1, column=1).value = f"更新日期: {today_str} (数据日期: {trade_day_str})"

    # 确保日报有触发信号列标题
    if ws_daily.cell(row=1, column=15).value is None:
        ws_daily.cell(row=1, column=15).value = "触发信号"

    # --- 2. 信号检测 ---
    # 补全 holdings_data 中的 weight 信息
    for i, hd in enumerate(holdings_data):
        row = CODE_ROWS[i]
        hd["weight"] = ws_summary.cell(row=row, column=6).value or 0
        hd["target_weight"] = ws_summary.cell(row=row, column=5).value or 0

    signals = detect_signals(holdings_data, trade_date_str)

    # 组合级别信号
    total_pnl_pct = (total_market_value - total_cost_value) / total_cost_value if total_cost_value > 0 else 0
    equity_weight = sum(
        hd["weight"] for hd in holdings_data if HOLDING_MAP[hd["code"]]["type"] == "equity"
    )
    portfolio_signals = detect_portfolio_signals(total_pnl_pct, equity_weight)

    # 写入备注（始终更新，无信号则清空）
    for i, hd in enumerate(holdings_data):
        row = CODE_ROWS[i]
        ss = signals.get(hd["code"], [])
        note = " | ".join(ss) if ss else ""
        ws_summary.cell(row=row, column=14).value = note if note else None

    # --- 3. 追加日报 sheet ---
    # 日报列结构: A日期 B品种 C代码 D持仓股数 E成本价 F持仓成本 G今日收盘 H涨跌幅 I当前市值 J浮动盈亏 K盈亏% L目标权重 M实际权重 N偏离 O触发信号
    # 检查是否已存在当日数据
    today_short = trade_date_str  # YYYYMMDD
    existing_rows_to_delete = []
    for r in range(2, ws_daily.max_row + 1):
        dval_norm = normalize_date_str(ws_daily.cell(row=r, column=1).value)
        if dval_norm == today_short:
            existing_rows_to_delete.append(r)

    if existing_rows_to_delete:
        # 删除已有当日数据行（从后往前删，避免行号漂移）
        print(f"  [INFO] 删除已有 {trade_date_str} 数据 {len(existing_rows_to_delete)} 行")
        for r in reversed(existing_rows_to_delete):
            ws_daily.delete_rows(r)

    # 追加新数据
    next_row = ws_daily.max_row + 1
    for hd in holdings_data:
        code = hd["code"]
        close = hd["close"]
        cost_price = hd["cost_price"]
        quantity = hd["quantity"]
        pct_chg = hd.get("pct_chg", 0)  # 原始百分比值，如 1.41 表示 +1.41%
        pnl = round((close - cost_price) * quantity, 2) if (quantity and cost_price) else 0
        pnl_pct = round((close - cost_price) / cost_price, 4) if (quantity and cost_price and cost_price > 0) else 0
        market_value = round(close * quantity, 2) if quantity else 0
        cost_value = round(cost_price * quantity, 2) if (cost_price and quantity) else 0
        ss = signals.get(code, [])
        signal_str = " | ".join(ss) if ss else "—"
        deviation = hd.get("weight", 0) - hd.get("target_weight", 0)
        target_weight = hd.get("target_weight", 0)

        ws_daily.cell(row=next_row, column=1).value = today_short         # A: 日期
        ws_daily.cell(row=next_row, column=2).value = hd["name"]          # B: 品种
        ws_daily.cell(row=next_row, column=3).value = str(code)           # C: 代码
        ws_daily.cell(row=next_row, column=4).value = quantity if quantity else None  # D: 持仓股数
        ws_daily.cell(row=next_row, column=5).value = cost_price if cost_price else None  # E: 成本价
        ws_daily.cell(row=next_row, column=6).value = cost_value if cost_value else None  # F: 持仓成本
        ws_daily.cell(row=next_row, column=7).value = close                # G: 今日收盘
        ws_daily.cell(row=next_row, column=8).value = round(pct_chg / 100, 4) if pct_chg else None  # H: 涨跌幅(小数)
        ws_daily.cell(row=next_row, column=9).value = market_value         # I: 当前市值
        ws_daily.cell(row=next_row, column=10).value = pnl                 # J: 浮动盈亏
        ws_daily.cell(row=next_row, column=11).value = pnl_pct             # K: 盈亏%
        ws_daily.cell(row=next_row, column=12).value = target_weight       # L: 目标权重
        ws_daily.cell(row=next_row, column=13).value = hd.get("weight")    # M: 实际权重
        ws_daily.cell(row=next_row, column=14).value = round(deviation, 4) # N: 偏离
        ws_daily.cell(row=next_row, column=15).value = signal_str          # O: 触发信号
        next_row += 1

    print(f"  [OK] 日报追加 {len(holdings_data)} 条记录 ({trade_date_str})")

    # --- 3.5 回补缺失的交易日日报数据 ---
    backfill_missing_trading_days(wb, ws_summary, ws_daily, trade_date_str)

    # --- 4. 保存 ---
    wb.save(FILE_PATH)
    print(f"  [OK] 文件已保存: {FILE_PATH}")
    return True


# ==== 仪表盘 HTML 同步 ====
import re

DASHBOARD_PATH = Path(r'C:\Users\65004\Desktop\小白\cc-data\portfolio-dashboard.html')

# 仪表盘显示用元数据（名称、分类、颜色）
DASHBOARD_META = {
    510300: {"displayName": "华泰柏瑞沪深300ETF", "targetWeight": 18.75, "cat": "沪深300", "catColor": "#60a5fa"},
    588050: {"displayName": "科创50ETF",          "targetWeight": 6.25,  "cat": "科创50",  "catColor": "#c084fc"},
    159915: {"displayName": "易方达创业板ETF",    "targetWeight": 6.25,  "cat": "创业板",  "catColor": "#e879f9"},
    512890: {"displayName": "华泰柏瑞红利低波ETF","targetWeight": 6.25,  "cat": "红利低波","catColor": "#f472b6"},
    511380: {"displayName": "可转债ETF",          "targetWeight": 12.5,  "cat": "可转债",  "catColor": "#fb923c"},
    511260: {"displayName": "国泰上证10年期国债ETF","targetWeight":12.5, "cat": "10年国债","catColor": "#34d399"},
    511010: {"displayName": "国泰上证5年期国债ETF","targetWeight": 12.5, "cat": "5年国债", "catColor": "#2dd4bf"},
    511360: {"displayName": "海富通中证短融ETF",  "targetWeight": 6.25,  "cat": "短融",    "catColor": "#22d3ee"},
    518880: {"displayName": "华安黄金ETF",        "targetWeight": 18.75, "cat": "黄金",    "catColor": "#fbbf24"},
}


def sync_dashboard():
    """
    从 Excel 读取最新持仓/流水数据，同步写入 portfolio-dashboard.html
    确保仪表盘网页的 basePositions / transactions 数组与 Excel 一致。
    """
    from openpyxl import load_workbook

    if not FILE_PATH.exists():
        print("  [WARN] Excel 不存在，跳过仪表盘同步")
        return
    if not DASHBOARD_PATH.exists():
        print("  [WARN] 仪表盘 HTML 不存在，跳过同步")
        return

    print("  正在同步仪表盘 HTML...")
    wb = load_workbook(FILE_PATH, data_only=True)
    ws_sum = wb["汇总"]

    # --- 1. 读取汇总表持仓数据 ---
    positions = {}
    for i, h in enumerate(HOLDINGS):
        row = CODE_ROWS[i]
        code = h["code"]
        cost_price = ws_sum.cell(row=row, column=8).value   # H: 成本价
        quantity = ws_sum.cell(row=row, column=9).value or 0  # I: 数量
        close_price = ws_sum.cell(row=row, column=11).value   # K: 今日收盘价
        cost_value = round(cost_price * quantity, 2) if cost_price and quantity else 0

        meta = DASHBOARD_META.get(code, {})
        positions[code] = {
            "cost_price": cost_price,
            "close_price": close_price,
            "quantity": quantity,
            "cost_value": cost_value,
            "displayName": meta.get("displayName", h["name"]),
            "targetWeight": meta.get("targetWeight", 0),
            "cat": meta.get("cat", ""),
            "catColor": meta.get("catColor", "#888"),
        }

    # 按 HOLDINGS 顺序生成 basePositions JS
    pos_lines = []
    for h in HOLDINGS:
        p = positions[h["code"]]
        cp = p["cost_price"]
        cp_str = f"{cp}" if cp else "null"
        clp = p.get("close_price")
        clp_str = f"{clp}" if clp else "null"
        line = (
            f"  {{ code:'{h['code']}', name:'{p['displayName']}', "
            f"targetWeight:{p['targetWeight']}, costPrice:{cp_str}, "
            f"closePrice:{clp_str}, qty:{p['quantity']}, cost:{p['cost_value']}, "
            f"cat:'{p['cat']}', catColor:'{p['catColor']}' }}"
        )
        pos_lines.append(line)

    new_positions_block = "var basePositions = [\n" + ",\n".join(pos_lines) + "\n];"

    # --- 2. 读取流水表 ---
    ws_trade = wb["流水"]
    txns = []
    for row in ws_trade.iter_rows(min_row=2, values_only=True):
        date_val, code, name, price, qty, amount = row[:6]
        if not date_val or not code:
            continue
        # 处理日期格式
        if isinstance(date_val, (date, datetime)):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            date_str = str(date_val)[:10]
        code_str = str(int(code)) if isinstance(code, float) else str(code)
        name_str = str(name) if name else ""
        price_val = float(price) if price else 0
        qty_val = int(qty) if qty else 0
        amount_val = float(amount) if amount else 0
        # 容错：公式缓存丢失时，用 price * abs(qty) 计算
        if not amount_val and price_val and qty_val:
            amount_val = round(price_val * abs(qty_val), 2)
        txns.append({
            "date": date_str, "code": code_str, "name": name_str,
            "price": price_val, "qty": qty_val, "amount": amount_val,
        })

    txn_lines = []
    for t in txns:
        line = (
            f"  {{ date:'{t['date']}', code:'{t['code']}', name:'{t['name']}', "
            f"price:{t['price']}, qty:{t['qty']}, amount:{t['amount']} }}"
        )
        txn_lines.append(line)

    new_txns_block = "var transactions = [\n" + ",\n".join(txn_lines) + "\n];"

    # --- 3. 读取日报表，汇总每日浮动盈亏 ---
    # 注意：历史数据列映射不一致，只有 9 行/天 且为新日期（8位数字格式）才可靠
    ws_daily = wb["日报"]
    raw_rows = {}  # date_str -> list of row tuples
    for row in ws_daily.iter_rows(min_row=2, values_only=True):
        date_val = row[0]
        if not date_val:
            continue
        # 保留原始字符串用于判断格式
        if isinstance(date_val, (date, datetime)):
            date_str = date_val.strftime("%Y-%m-%d")
            raw_fmt = "date"  # 日期对象，旧格式
        else:
            raw = str(date_val).strip()
            if len(raw) == 8 and raw.isdigit():
                date_str = f"{raw[:4]}-{raw[4:6]}-{raw[6:8]}"
                raw_fmt = "8digit"  # 当前脚本格式
            else:
                date_str = raw[:10] if len(raw) >= 10 else raw
                raw_fmt = "string"
        if date_str not in raw_rows:
            raw_rows[date_str] = {"rows": [], "fmt": raw_fmt}
        raw_rows[date_str]["rows"].append(row)

    # 只使用 9行/天 + 8位数字日期格式 + P&L在合理范围内的数据
    # BUGFIX(v2026.06.27): 过滤早期列映射不一致的数据（P&L异常大=市场价值）
    # BUGFIX(v2026.06.27): 额外检查是否至少有品种持仓>0（排除qty=0的垃圾行）
    daily_totals = {}
    for date_str, info in raw_rows.items():
        if len(info["rows"]) == 9 and info["fmt"] == "8digit":
            total_pnl = sum(float(r[9]) if r[9] else 0 for r in info["rows"])
            # 合理的总P&L范围：±500k（4M组合的±12.5%）
            if abs(total_pnl) <= 500000:
                # 进一步检查：至少有一个品种持仓数量>0且收盘价有效，或有浮动盈亏数据（排除垃圾行）
                # v2026.07.02-r8: 部分日期（如6/30）修复后qty列为空但close/pnl有效，需放宽条件
                has_real_data = any(
                    ((r[3] or 0) > 0 and (r[6] or 0) > 0) or (r[9] is not None)
                    for r in info["rows"]
                )
                if has_real_data:
                    daily_totals[date_str] = round(total_pnl, 2)

    # --- 3.5 读取流水原始数据（用于已实现利润计算，必须在 wb.close() 前） ---
    txn_raw = []  # [(date, code, name, price, qty), ...]
    for row in ws_trade.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            d_val = row[0]
            if isinstance(d_val, (date, datetime)):
                d_str = d_val.strftime("%Y-%m-%d")
            else:
                d_str = str(d_val)[:10]
            txn_raw.append((
                d_str,
                str(int(row[1])) if isinstance(row[1], float) else str(row[1]),
                str(row[2]) if row[2] else '',
                float(row[3]) if row[3] else 0,
                int(row[4]) if row[4] else 0
            ))

    wb.close()

    # --- 4. 读取并更新 HTML ---
    html = DASHBOARD_PATH.read_text(encoding="utf-8")

    html = re.sub(
        r'var basePositions = \[[\s\S]*?\n\];',
        new_positions_block,
        html,
        count=1,
    )
    html = re.sub(
        r'var transactions = \[[\s\S]*?\n\];',
        new_txns_block,
        html,
        count=1,
    )

    # --- 5. 解析现有 SEED_DAILY（用于后续修正对比和桥接） ---
    existing_entries = []  # [(date, pnl, cumulative), ...]
    seed_match = re.search(r'var SEED_DAILY = \[([\s\S]*?)\n\];', html)
    if seed_match:
        entry_pattern = re.compile(r"date:'(\d{4}-\d{2}-\d{2})',\s*pnl:([-\d.]+),\s*cumulative:([-\d.]+)")
        for m in entry_pattern.finditer(seed_match.group(1)):
            existing_entries.append((m.group(1), float(m.group(2)), float(m.group(3))))

    # --- 6. SEED_DAILY 更新：将日报总P&L转为每日变化量 ---
    # BUGFIX(v2026.06.27): 日报列J是累计浮动盈亏，需取相邻日差值得每日变化
    # v2026.07.02-r8: clean dates 的 cumulative 直接用日报 col J 总和，
    # 不再从 SEED_DAILY 桥接（避免早期累计偏移传导到后续所有日期）
    sorted_clean_dates = sorted(daily_totals.keys())
    daily_changes = {}
    prev_total = None
    for d in sorted_clean_dates:
        curr_total = daily_totals[d]
        if prev_total is not None:
            daily_changes[d] = round(curr_total - prev_total, 2)
        else:
            daily_changes[d] = 0.0  # 首个日期占位，后续会被覆盖
        prev_total = curr_total

    # v2026.07.02-r7: 日报 6/19 之后数据已清洁，以日报为权威覆盖已有条目
    # 6/18 及之前的数据仍可能有损坏，保留现有 SEED_DAILY 不受日报影响
    CORRUPTION_CUTOFF = '2026-06-19'
    existing_date_set = {e[0] for e in existing_entries}

    # 日报中有、但 SEED_DAILY 没有的 → 新增（直接使用日报累计值）
    all_raw = []
    for d in sorted_clean_dates:
        if d not in existing_date_set:
            all_raw.append((d, daily_changes[d], daily_totals[d]))

    # 日报和 SEED_DAILY 都有的 → 6/19之后用日报覆盖（pnl + cumulative），6/18之前保留 SEED_DAILY
    for d, pnl, cum in existing_entries:
        if d in sorted_clean_dates and d >= CORRUPTION_CUTOFF:
            all_raw.append((d, daily_changes[d], daily_totals[d]))  # 日报覆盖 pnl + cumulative
        else:
            all_raw.append((d, pnl, cum))  # 保留现有 pnl + cumulative

    all_raw.sort(key=lambda x: x[0])

    # 过滤非交易日，构建最终列表
    # v2026.07.02-r8: clean dates 直接用日报 cumulative；early dates 用原有 cumulative
    # pnl 统一从相邻日期的 cumulative 差值计算，确保内部一致
    all_entries = []
    for d_str, day_pnl, target_cum in all_raw:
        entry_date = date(int(d_str[:4]), int(d_str[5:7]), int(d_str[8:10]))
        if not is_trading_day(entry_date):
            continue
        all_entries.append([d_str, day_pnl, target_cum])  # 暂存 target_cum

    # 重新计算 pnl = cumulative 差值（保证过渡日期一致）
    for i in range(len(all_entries)):
        d_str, _, target_cum = all_entries[i]
        if i == 0:
            all_entries[i][1] = 0.0  # 首个日期 pnl 置 0（无前一天可比较）
        else:
            prev_cum = all_entries[i-1][2]
            all_entries[i][1] = round(target_cum - prev_cum, 2)
        all_entries[i][2] = round(target_cum, 2)

    # 转换回 tuple 列表
    all_entries = [(e[0], e[1], e[2]) for e in all_entries]

    # 日志
    added_dates = [d for d in sorted_clean_dates if d not in existing_date_set]
    if added_dates:
        print(f"  [INFO] SEED_DAILY 追加 {len(added_dates)} 个新日期: {added_dates}")
    else:
        print(f"  [INFO] SEED_DAILY 无更新")

    seed_lines = []
    for date_str, pnl, cum in all_entries:
        seed_lines.append(f"  {{ date:'{date_str}', pnl:{pnl}, cumulative:{cum} }}")
    new_seed_block = "var SEED_DAILY = [\n" + ",\n".join(seed_lines) + "\n];"

    html = re.sub(
        r'var SEED_DAILY = \[[\s\S]*?\n\];',
        new_seed_block,
        html,
        count=1,
    )

    # --- 7. 计算已实现利润（FIFO 匹配买入成本） ---
    txn_data = sorted(txn_raw, key=lambda x: x[0])

    realized = []
    buy_queues = {}  # code -> [[date, price, qty], ...]
    for t_date, t_code, t_name, t_price, t_qty in txn_data:
        if t_qty > 0:
            if t_code not in buy_queues:
                buy_queues[t_code] = []
            buy_queues[t_code].append([t_date, t_price, t_qty])
        elif t_qty < 0:
            sell_qty = -t_qty
            sell_price = t_price
            if t_code in buy_queues and buy_queues[t_code]:
                remaining = sell_qty
                total_profit = 0.0
                matched_cost_price = 0.0
                while remaining > 0 and buy_queues[t_code]:
                    lot = buy_queues[t_code][0]
                    take = min(remaining, lot[2])
                    profit_on_lot = (sell_price - lot[1]) * take
                    total_profit += profit_on_lot
                    matched_cost_price = lot[1]
                    lot[2] -= take
                    remaining -= take
                    if lot[2] <= 0:
                        buy_queues[t_code].pop(0)
                realized.append({
                    'date': t_date,
                    'code': t_code,
                    'name': t_name,
                    'sellPrice': round(sell_price, 4),
                    'costPrice': round(matched_cost_price, 4),
                    'qty': sell_qty,
                    'profit': round(total_profit, 2)
                })

    # 生成 realizedProfits 代码块
    if realized:
        rp_lines = []
        for r in realized:
            rp_lines.append(
                f"  {{ date:'{r['date']}', code:'{r['code']}', name:'{r['name']}', "
                f"sellPrice:{r['sellPrice']}, costPrice:{r['costPrice']}, qty:{r['qty']}, profit:{r['profit']} }}"
            )
        new_rp_block = "var realizedProfits = [\n" + ",\n".join(rp_lines) + "\n];"
        html = re.sub(
            r'var realizedProfits = \[[\s\S]*?\n\];',
            new_rp_block,
            html,
            count=1,
        )

    # dividends 保留 HTML 中已有条目，不覆盖

    DASHBOARD_PATH.write_text(html, encoding="utf-8")
    print(f"  [OK] 仪表盘已同步 ({len(pos_lines)} 持仓, {len(txn_lines)} 笔交易, {len(all_entries)} 个日报快照, {len(realized)} 笔已实现利润)")


# ==== 主流程 ====
def main():
    print(f"\n{'='*60}")
    print(f"  400万资产配置组合 - 每日更新")
    print(f"  执行时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*60}")

    # 1. 判断交易日
    today = date.today()
    trade_day = get_latest_trading_day()
    trade_date_str = trade_day.strftime("%Y%m%d")

    if not is_trading_day(today):
        print(f"  [SKIP] {today} 非交易日，使用最近交易日 {trade_day}")
        # 如果是非交易日但想更新最近交易日数据，可以继续
        # return  # 如需严格跳过，取消此行注释

    print(f"  交易日: {trade_day} ({trade_date_str})")

    # 2. 检查文件
    if not FILE_PATH.exists():
        print(f"  [ERROR] 文件不存在: {FILE_PATH}")
        return 1

    # 3. 获取行情
    print("  正在获取行情数据...")
    ts_codes = [h["ts"] for h in HOLDINGS]
    prices = fetch_prices(ts_codes, trade_date_str)

    if not prices:
        print("  [ERROR] 未能获取任何行情数据，请检查 Tushare 连接和权限")
        return 1

    print(f"  获取到 {len(prices)}/{len(ts_codes)} 个品种价格")
    for ts_code, info in prices.items():
        print(f"    {ts_code}: close={info['close']}, chg={info['pct_chg']:+.2f}%")

    # 4. 更新 Excel
    print("  正在更新表格...")
    try:
        update_spreadsheet(prices, trade_date_str)
    except Exception as e:
        print(f"  [ERROR] 更新失败: {e}")
        import traceback
        traceback.print_exc()
        return 1

    # 5. 同步仪表盘 HTML
    try:
        sync_dashboard()
    except Exception as e:
        print(f"  [WARN] 仪表盘同步失败（不中断主流程）: {e}")

    # 6. 保存缓存（记录最后更新时间）
    cache = {"last_update": datetime.now().isoformat(), "trade_date": trade_date_str}
    try:
        with open(CACHE_PATH, "wb") as f:
            pickle.dump(cache, f)
    except Exception:
        pass

    print(f"\n  [OK] 更新完成!")
    return 0


if __name__ == "__main__":
    sys.exit(main())
