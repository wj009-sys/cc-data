#!/usr/bin/env python
"""
400万资产配置组合 - 投资简报生成与微信推送
==========================================
支持两种模式:
  midday     午间简报 (11:45): 东方财富实时行情 + Excel持仓数据
  afternoon  收盘简报 (16:30): 读取已更新Excel完整数据
推送渠道: 企业微信群机器人

用法:
  python generate_briefing.py midday       # 生成并推送午间简报
  python generate_briefing.py afternoon    # 生成并推送收盘简报
"""
import os
import sys
import io
import json
import requests
from datetime import datetime, date, timedelta
from pathlib import Path

# 修复 Windows GBK
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# ==== 配置 ====
FILE_PATH = Path(r'C:\Users\65004\Desktop\小白\cc-data\400万资产配置组合2026.xlsx')
WECOM_WEBHOOK = os.environ.get('WECOM_WEBHOOK') or 'https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=9cccf118-070c-422f-ada6-1a8db2e0f309'

# 9 个 ETF 持仓
HOLDINGS = [
    {"code": "510300", "name": "沪深300ETF",   "short": "沪深300",  "type": "equity"},
    {"code": "588050", "name": "科创50ETF",    "short": "科创50",   "type": "equity", "high_vol": True},
    {"code": "159915", "name": "创业板ETF",    "short": "创业板",   "type": "equity", "high_vol": True},
    {"code": "512890", "name": "红利低波ETF",  "short": "红利低波", "type": "equity"},
    {"code": "511380", "name": "可转债ETF",    "short": "可转债",   "type": "equity"},
    {"code": "511260", "name": "10年国债ETF",  "short": "10年国债", "type": "bond"},
    {"code": "511010", "name": "5年国债ETF",   "short": "5年国债",  "type": "bond"},
    {"code": "511360", "name": "短融ETF",      "short": "短融",     "type": "bond"},
    {"code": "518880", "name": "黄金ETF",      "short": "黄金",     "type": "commodity"},
]
CODE_ROWS = [2, 3, 4, 5, 6, 7, 8, 9, 10]
HOLDING_MAP = {h["code"]: h for h in HOLDINGS}

# 参考指数（用于大盘参考）
REF_INDICES = [
    {"code": "000300", "name": "沪深300", "secid": "1.000300"},
    {"code": "000688", "name": "科创50",  "secid": "1.000688"},
    {"code": "399006", "name": "创业板指","secid": "0.399006"},
]


# ============================================================
#  工具函数
# ============================================================

def get_secid(code: str) -> str:
    """东方财富 secid: 上海代码前缀1，深圳代码前缀0"""
    c = str(code).strip()
    if c.startswith("159") or c.startswith("0") or c.startswith("3"):
        return f"0.{c}"
    return f"1.{c}"


def fmt(val, suffix="%"):
    """格式化涨跌幅/盈亏百分比"""
    if val is None:
        return "N/A"
    if val > 0:
        return f"+{val:.2f}{suffix}"
    return f"{val:.2f}{suffix}"


def fmt_money(val):
    """格式化金额，单位自动切换"""
    if val is None:
        return "N/A"
    if abs(val) >= 10000:
        return f"{val/10000:.2f}万"
    return f"{val:.0f}元"


def fmt_pnl(val):
    """格式化盈亏数值（带正负号）"""
    if val is None:
        return "N/A"
    if val >= 0:
        return f"+{val:.0f}"
    return f"{val:.0f}"


# ============================================================
#  东方财富实时行情
# ============================================================

def fetch_realtime(codes: list) -> dict:
    """
    从东方财富批量获取实时行情
    返回 {code: {price, change_pct, name}}
    """
    secids = [get_secid(c) for c in codes]
    url = (
        "https://push2.eastmoney.com/api/qt/ulist.np/get"
        f"?fltt=2&fields=f2,f3,f4,f12,f14"
        f"&secids={','.join(secids)}"
        f"&_={int(datetime.now().timestamp() * 1000)}"
    )
    try:
        resp = requests.get(url, timeout=10, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Referer": "https://quote.eastmoney.com/",
        })
        data = resp.json()
        if not data.get("data") or not data["data"].get("diff"):
            print("  [WARN] 东方财富返回格式异常:", json.dumps(data, ensure_ascii=False)[:200])
            return {}

        result = {}
        for item in data["data"]["diff"]:
            code = str(item.get("f12", ""))
            result[code] = {
                "price": item.get("f2"),
                "change_pct": item.get("f3"),
                "change_amt": item.get("f4"),
                "name": item.get("f14", ""),
            }
        return result
    except requests.exceptions.Timeout:
        print("  [ERROR] 东方财富请求超时")
    except requests.exceptions.ConnectionError:
        print("  [ERROR] 东方财富连接失败")
    except Exception as e:
        print(f"  [ERROR] 东方财富行情异常: {e}")
    return {}


# ============================================================
#  Excel 数据读取
# ============================================================

def read_excel_summary() -> dict:
    """
    从汇总 sheet 读取持仓数据
    返回 {total_cost, total_mv, total_pnl, total_pnl_pct, positions: [{...}]}
    """
    from openpyxl import load_workbook

    wb = load_workbook(FILE_PATH, data_only=True)
    ws = wb["汇总"]

    positions = []
    total_cost = 0.0
    total_mv = 0.0

    for i, h in enumerate(HOLDINGS):
        row = CODE_ROWS[i]

        def get(col):
            v = ws.cell(row=row, column=col).value
            return v if v else 0

        code = h["code"]
        cost_price = get(8)    # H: 成本价
        quantity = get(9)      # I: 数量
        target_w = get(5)      # E: 目标权重
        last_close = get(11)   # K: 最新收盘价
        pnl = get(12)          # L: 浮动盈亏
        pnl_pct = get(13)      # M: 盈亏%
        act_weight = get(6)    # F: 当前权重
        deviation = get(7)     # G: 偏离
        signal_raw = ws.cell(row=row, column=14).value  # N: 信号
        signal = str(signal_raw).strip() if signal_raw else ""

        cost_v = cost_price * quantity if cost_price and quantity else 0
        mv = last_close * quantity if last_close and quantity else 0
        total_cost += cost_v
        total_mv += mv

        positions.append({
            "code": code, "name": h["name"], "short": h["short"],
            "type": h["type"], "high_vol": h.get("high_vol", False),
            "cost_price": cost_price, "quantity": quantity,
            "target_weight": target_w, "last_close": last_close,
            "pnl": pnl, "pnl_pct": pnl_pct,
            "actual_weight": act_weight, "deviation": deviation,
            "signal": signal,
            "cost_value": cost_v, "market_value": mv,
        })

    wb.close()

    total_pnl = total_mv - total_cost
    return {
        "total_cost": total_cost,
        "total_mv": total_mv,
        "total_pnl": total_pnl,
        "total_pnl_pct": total_pnl / total_cost if total_cost > 0 else 0,
        "positions": positions,
    }


def read_previous_trading_day_mv(today_str: str) -> float:
    """
    从日报 sheet 读取最近一个非今日交易日的总市值（用于计算今日变动）
    today_str: YYYYMMDD 格式的今天
    返回前一日总市值（元），0 表示无法获取
    """
    from openpyxl import load_workbook
    try:
        wb = load_workbook(FILE_PATH, data_only=True)
        ws = wb["日报"]
        date_groups = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            dval = row[0]
            col9 = row[8]  # I: 当前市值
            if not dval or col9 is None:
                continue
            if isinstance(dval, (date, datetime)):
                ds = dval.strftime("%Y%m%d")
            else:
                raw = str(dval).strip()
                if len(raw) == 8 and raw.isdigit():
                    ds = raw
                elif len(raw) >= 10:
                    ds = raw[:10].replace("-", "")
                else:
                    continue
            if ds not in date_groups:
                date_groups[ds] = []
            date_groups[ds].append(float(col9))

        # 找最近的非今日交易日
        sorted_dates = sorted(d for d in date_groups if len(date_groups[d]) == 9)
        for d in reversed(sorted_dates):
            if d != today_str:
                wb.close()
                return sum(date_groups[d])
        wb.close()
        return 0
    except Exception as e:
        print(f"  [WARN] 读取日报历史市值失败: {e}")
        return 0


def read_today_daily_changes(today_str: str) -> dict:
    """
    从日报 sheet 读取今日涨跌幅数据
    返回 {code: {close, pct_chg, pnl, pnl_pct}}
    """
    from openpyxl import load_workbook
    try:
        wb = load_workbook(FILE_PATH, data_only=True)
        ws = wb["日报"]
        code_col = {}  # code -> row data
        for row in range(2, ws.max_row + 1):
            dval = ws.cell(row=row, column=1).value
            if not dval:
                continue
            if isinstance(dval, (date, datetime)):
                ds = dval.strftime("%Y%m%d")
            else:
                raw = str(dval).strip()
                if len(raw) == 8 and raw.isdigit():
                    ds = raw
                elif len(raw) >= 10:
                    ds = raw[:10].replace("-", "")
                else:
                    continue
            if ds != today_str:
                continue
            code_raw = ws.cell(row=row, column=3).value
            if code_raw is None:
                continue
            c = str(int(code_raw)) if isinstance(code_raw, float) else str(code_raw).strip()
            code_col[c] = {
                "close": ws.cell(row=row, column=7).value,   # G: 收盘
                "pct_chg": ws.cell(row=row, column=8).value, # H: 涨跌幅
                "pnl": ws.cell(row=row, column=10).value,    # J: 浮动盈亏
                "pnl_pct": ws.cell(row=row, column=11).value, # K: 盈亏%
            }
        wb.close()
        return code_col
    except Exception as e:
        print(f"  [WARN] 读取日报今日数据失败: {e}")
        return {}


# ============================================================
#  简报生成
# ============================================================

def generate_midday_briefing() -> str:
    """
    午间简报 (11:45)
    使用东方财富实时行情 + Excel持仓数据
    """
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    today_short = now.strftime("%Y%m%d")
    weekday_cn = ["一", "二", "三", "四", "五", "六", "日"][now.weekday()]

    print("  [1/3] 读取 Excel 持仓数据...")
    portfolio = read_excel_summary()
    positions = portfolio["positions"]

    print("  [2/3] 获取实时行情...")
    all_codes = [h["code"] for h in HOLDINGS] + [idx["code"] for idx in REF_INDICES]
    realtime = fetch_realtime(all_codes)

    # 检查实时行情是否获取到关键ETF数据
    has_realtime = any(c in realtime and realtime[c].get("price") is not None
                       for c in [h["code"] for h in HOLDINGS])

    lines = []
    lines.append(f"📊 **400万投资组合 · 午间简报**")
    lines.append(f"📅 {date_str} 周{weekday_cn} 11:45")
    lines.append("")

    if not has_realtime:
        lines.append("⚠️ 实时行情暂不可用，以下为收盘数据参考")
        lines.append("")

    # --- 大盘参考 ---
    idx_parts = []
    for idx in REF_INDICES:
        p = realtime.get(idx["code"])
        if p and p.get("change_pct") is not None:
            idx_parts.append(f"{idx['name']} {fmt(p['change_pct'])}")
    if idx_parts:
        lines.append("**大盘参考**")
        lines.append("> " + " | ".join(idx_parts))
        lines.append("")

    # --- 组合概览 ---
    if has_realtime:
        # 用实时价格计算当前市值
        today_mv = 0.0
        today_pnl_sum = 0.0
        items = []

        for p in positions:
            code = p["code"]
            rt = realtime.get(code, {})
            cur_price = rt.get("price") if rt.get("price") is not None else p["last_close"]
            qty = p["quantity"]
            if qty > 0:
                cur_mv = cur_price * qty
                day_pnl = (cur_price - p["last_close"]) * qty if p["last_close"] > 0 else 0
                total_pnl_val = (cur_price - p["cost_price"]) * qty if p["cost_price"] > 0 else 0
                total_pnl_pct_val = (cur_price - p["cost_price"]) / p["cost_price"] if p["cost_price"] > 0 else 0
                day_pct = rt.get("change_pct")  # 今日涨跌%
            else:
                cur_mv = 0
                day_pnl = 0
                total_pnl_val = 0
                total_pnl_pct_val = 0
                day_pct = 0

            today_mv += cur_mv
            today_pnl_sum += day_pnl
            items.append({**p, "cur_price": cur_price, "cur_mv": cur_mv,
                          "day_pnl": day_pnl, "day_pct": day_pct,
                          "total_pnl_val": total_pnl_val,
                          "total_pnl_pct_val": total_pnl_pct_val})

        yesterday_mv = read_previous_trading_day_mv(today_short) or portfolio["total_mv"]
        if yesterday_mv > 0:
            day_pnl_pct = today_pnl_sum / yesterday_mv * 100
        else:
            day_pnl_pct = 0

        total_return = (today_mv - portfolio["total_cost"]) / portfolio["total_cost"] * 100 if portfolio["total_cost"] > 0 else 0

        lines.append(f"**组合概览**")
        lines.append(f"总市值: {fmt_money(today_mv)}")
        lines.append(f"今日: {fmt_pnl(today_pnl_sum)} ({fmt(day_pnl_pct)})")
        lines.append(f"累计: {fmt_pnl(today_mv - portfolio['total_cost'])} ({fmt(total_return)})")
        lines.append("")

        # --- 今日涨跌排序 ---
        valid_items = [i for i in items if i["day_pct"] is not None]
        if valid_items:
            sorted_items = sorted(valid_items, key=lambda x: x["day_pct"], reverse=True)
            lines.append("**今日表现**")
            for i in sorted_items:
                emoji = "🟢" if i["day_pct"] >= 0 else "🔴"
                pnl_e = "🟢" if i["total_pnl_pct_val"] >= 0 else "🔴"
                pnl_str = fmt(i["total_pnl_pct_val"] * 100)
                lines.append(f"{emoji}{i['short']} {fmt(i['day_pct'])} {pnl_e}{pnl_str}")
            lines.append("")

        # --- 偏离最大的品种 ---
        dev_items = [i for i in items if i["target_weight"] > 0 and i["cur_mv"] > 0]
        if dev_items:
            for i in dev_items:
                i["cur_weight"] = i["cur_mv"] / today_mv if today_mv > 0 else 0
                i["cur_deviation"] = i["cur_weight"] - i["target_weight"]
            max_dev = max(dev_items, key=lambda x: abs(x["cur_deviation"]))
            if abs(max_dev["cur_deviation"]) > 0.01:
                dir_str = "超配" if max_dev["cur_deviation"] > 0 else "低配"
                lines.append(f"**偏离提醒**")
                lines.append(f">{max_dev['short']} {dir_str} {abs(max_dev['cur_deviation'])*100:.1f}%")
                lines.append("")

    else:
        # 无实时行情，使用 Excel 数据
        lines.append(f"**组合概览**")
        lines.append(f"总市值: {fmt_money(portfolio['total_mv'])}")
        lines.append(f"累计盈亏: {fmt_pnl(portfolio['total_pnl'])} ({fmt(portfolio['total_pnl_pct']*100)})")
        lines.append("")

    # --- 信号提醒 ---
    active_signals = [p for p in positions if p["signal"] and p["signal"] not in ("—", "")]
    if active_signals:
        lines.append("**信号提醒**")
        for p in active_signals:
            # 截断过长的信号
            sig = p["signal"][:60]
            lines.append(f"⚠️ {p['short']}: {sig}")
        lines.append("")

    # --- 仓位分布 ---
    if has_realtime:
        eq_w = sum(i["cur_mv"] for i in items if i["type"] == "equity") / today_mv * 100 if today_mv > 0 else 0
        bd_w = sum(i["cur_mv"] for i in items if i["type"] == "bond") / today_mv * 100 if today_mv > 0 else 0
        gd_w = sum(i["cur_mv"] for i in items if i["type"] == "commodity") / today_mv * 100 if today_mv > 0 else 0
    else:
        eq_w = sum(p["market_value"] for p in positions if p["type"] == "equity") / portfolio["total_mv"] * 100 if portfolio["total_mv"] > 0 else 0
        bd_w = sum(p["market_value"] for p in positions if p["type"] == "bond") / portfolio["total_mv"] * 100 if portfolio["total_mv"] > 0 else 0
        gd_w = sum(p["market_value"] for p in positions if p["type"] == "commodity") / portfolio["total_mv"] * 100 if portfolio["total_mv"] > 0 else 0
    cash_w = max(0, 100 - eq_w - bd_w - gd_w)

    lines.append("**仓位分布**")
    lines.append(f"权益 {eq_w:.1f}% | 固收 {bd_w:.1f}% | 黄金 {gd_w:.1f}% | 现金 {cash_w:.1f}%")

    return "\n".join(lines)


def generate_afternoon_briefing() -> str:
    """
    收盘简报 (16:30)
    读取 update_portfolio.py 已更新的 Excel 数据
    """
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    today_short = now.strftime("%Y%m%d")
    weekday_cn = ["一", "二", "三", "四", "五", "六", "日"][now.weekday()]

    print("  [1/3] 读取 Excel 汇总数据...")
    portfolio = read_excel_summary()
    positions = portfolio["positions"]

    print("  [2/3] 读取今日涨跌幅明细...")
    daily_chg = read_today_daily_changes(today_short)

    # 有今日日报数据时，用它补充每日涨跌
    if daily_chg:
        for p in positions:
            dc = daily_chg.get(p["code"])
            if dc:
                if dc.get("pct_chg") is not None:
                    p["daily_pct"] = dc["pct_chg"] * 100  # 转百分比
                if dc.get("close") is not None:
                    p["last_close"] = dc["close"]

    total_mv = portfolio["total_mv"]
    total_cost = portfolio["total_cost"]
    total_pnl = portfolio["total_pnl"]
    total_pnl_pct = portfolio["total_pnl_pct"] * 100

    # 计算今日组合盈亏
    yesterday_mv = read_previous_trading_day_mv(today_short)
    if yesterday_mv > 0:
        today_pnl = total_mv - yesterday_mv
        today_pnl_pct = today_pnl / yesterday_mv * 100
    else:
        today_pnl = 0
        today_pnl_pct = 0

    # 切分类型
    eq_items = [p for p in positions if p["type"] == "equity"]
    bd_items = [p for p in positions if p["type"] == "bond"]
    gd_items = [p for p in positions if p["type"] == "commodity"]

    eq_mv = sum(p["market_value"] for p in eq_items)
    bd_mv = sum(p["market_value"] for p in bd_items)
    gd_mv = sum(p["market_value"] for p in gd_items)

    lines = []
    lines.append(f"📊 **400万投资组合 · 收盘简报**")
    lines.append(f"📅 {date_str} 周{weekday_cn} 16:30")
    lines.append("")

    # --- 组合概览 ---
    lines.append(f"**组合概览**")
    lines.append(f"总市值: {fmt_money(total_mv)}")
    lines.append(f"今日: {fmt_pnl(today_pnl)} ({fmt(today_pnl_pct)})")
    lines.append(f"累计盈亏: {fmt_pnl(total_pnl)} ({fmt(total_pnl_pct)})")
    lines.append(f"累计收益: {fmt(total_pnl_pct)}")
    lines.append("")

    # --- 品种明细 ---
    lines.append("**品种明细**")
    for p in positions:
        qty = int(p["quantity"])
        close = p["last_close"]
        if qty <= 0:
            continue
        # 涨跌幅
        if "daily_pct" in p:
            chg_str = fmt(p["daily_pct"])
        else:
            chg_str = f"{p['pnl_pct']*100:+.1f}%" if p["pnl_pct"] else "—"
        # 浮盈
        pnl_str = f"{p['pnl_pct']*100:+.1f}%" if p["pnl_pct"] else "—"
        # 偏离
        dev_str = f"{p['deviation']*100:+.1f}%" if p["deviation"] else "—"
        # 超配/低配标记
        if p["deviation"] > 0.005:
            dev_str += "↑"
        elif p["deviation"] < -0.005:
            dev_str += "↓"

        lines.append(f"> {p['short']}  {close}  {chg_str}  {pnl_str}  {dev_str}")
    lines.append("")

    # --- 信号提醒 ---
    active_signals = [p for p in positions if p["signal"] and p["signal"] not in ("—", "")]
    if active_signals:
        lines.append("**信号提醒**")
        for p in active_signals:
            sig = p["signal"][:80]
            lines.append(f"⚠️ **{p['short']}**: {sig}")
        lines.append("")

    # --- 仓位分布 ---
    eq_w = eq_mv / total_mv * 100 if total_mv > 0 else 0
    bd_w = bd_mv / total_mv * 100 if total_mv > 0 else 0
    gd_w = gd_mv / total_mv * 100 if total_mv > 0 else 0
    cash_w = max(0, 100 - eq_w - bd_w - gd_w)

    lines.append("**仓位分布**")
    lines.append(f"权益 {eq_w:.1f}% | 固收 {bd_w:.1f}% | 黄金 {gd_w:.1f}% | 现金 {cash_w:.1f}%")
    lines.append("")

    # --- 要点总结 ---
    lines.append("**📌 要点**")
    signal_count = len(active_signals)
    lines.append(f"调仓信号: {'有('+str(signal_count)+'条)' if signal_count > 0 else '无'}")

    # 偏离最大的品种
    max_dev_p = max(positions, key=lambda x: abs(x["deviation"]) if x["quantity"] > 0 else 0)
    if abs(max_dev_p["deviation"]) > 0.01 and max_dev_p["quantity"] > 0:
        dev_dir = "超配" if max_dev_p["deviation"] > 0 else "低配"
        lines.append(f"偏离最大: {max_dev_p['short']} {dev_dir} {abs(max_dev_p['deviation'])*100:.1f}%")

    # 持仓天数
    first_holdings = [p for p in positions if p["quantity"] > 0]
    lines.append(f"持仓品种: {len(first_holdings)} 个")

    return "\n".join(lines)


# ============================================================
#  企业微信发送
# ============================================================

def send_wecom(content: str) -> bool:
    """发送 markdown 消息到企业微信群机器人"""
    # 企业微信限制 4096 字节
    encoded = content.encode("utf-8")
    if len(encoded) > 4096:
        # 截断到 4000 字节的安全边界
        truncated = encoded[:4000]
        # 从最后一个换行处截断，避免断在半角字符中间
        last_newline = truncated.rfind(b"\n")
        if last_newline > 3000:
            truncated = truncated[:last_newline]
        content = truncated.decode("utf-8", errors="ignore") + "\n\n...（内容已截断）"
        print(f"  [WARN] 简报超长，已截断至 {len(content)} 字节")

    payload = {
        "msgtype": "markdown",
        "markdown": {"content": content},
    }

    try:
        resp = requests.post(
            WECOM_WEBHOOK,
            json=payload,
            timeout=15,
            headers={"Content-Type": "application/json"},
        )
        result = resp.json()
        if result.get("errcode") == 0:
            print(f"  [OK] 企业微信推送成功")
            return True
        else:
            print(f"  [ERROR] 企业微信错误: {result}")
            print(f"  [DEBUG] 内容长度: {len(content)} 字符 / {len(content.encode('utf-8'))} 字节")
            return False
    except requests.exceptions.Timeout:
        print("  [ERROR] 企业微信请求超时")
    except Exception as e:
        print(f"  [ERROR] 企业微信发送异常: {e}")

    return False


# ============================================================
#  主流程
# ============================================================

def main():
    if len(sys.argv) < 2:
        print("用法: python generate_briefing.py [midday|afternoon]")
        print("  midday      午间简报 (11:45，实时行情)")
        print("  afternoon   收盘简报 (16:30，Excel数据)")
        return 1

    mode = sys.argv[1].lower()
    mode_name = {"midday": "午间简报", "afternoon": "收盘简报"}
    label = mode_name.get(mode, mode)

    print(f"{'='*50}")
    print(f"  400万投资组合 - {label}")
    print(f"  执行时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*50}")

    # 检查 Excel
    if not FILE_PATH.exists():
        print(f"  [ERROR] Excel 文件不存在: {FILE_PATH}")
        return 1

    # 生成简报
    try:
        if mode == "midday":
            content = generate_midday_briefing()
        elif mode == "afternoon":
            content = generate_afternoon_briefing()
        else:
            print(f"  [ERROR] 未知模式: {mode}")
            return 1
    except Exception as e:
        print(f"  [ERROR] 生成简报失败: {e}")
        import traceback
        traceback.print_exc()
        return 1

    # 输出到控制台
    print(f"\n{'─'*45}")
    print(content)
    print(f"{'─'*45}")
    print(f"  简报长度: {len(content)} 字符 / {len(content.encode('utf-8'))} 字节")

    # 推送
    print(f"  正在推送到企业微信...")
    success = send_wecom(content)

    print(f"\n  {'[OK] 简报已发送' if success else '[WARN] 推送失败，请检查网络/webhook'}")
    return 0 if success else 1


if __name__ == "__main__":
    sys.exit(main())
