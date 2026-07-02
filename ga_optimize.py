#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
达尔文遗传算法优化器 — 投资组合规则参数优化
真正的 GA：锦标赛选择 + SBX 交叉 + 多项式突变 + 精英保留
种群 60 × 50 代 = 3000 次回测，优化 5 个连续参数
"""
import os, sys, json, math, time, random
from datetime import datetime, timedelta
import pandas as pd
import numpy as np
sys.stdout.reconfigure(encoding='utf-8')

import tushare as ts
ts.set_token(os.environ.get('TUSHARE_TOKEN', ''))
pro = ts.pro_api()

TOTAL = 4_000_000

# ========== ETF 配置 ==========
ETFS = [
    {'code':'510300','mkt':'SH','name':'沪深300ETF', 'target':0.1875, 'enable_sp':True, 'enable_sl':True, 'is_bond':False},
    {'code':'588050','mkt':'SH','name':'科创50ETF',  'target':0.0625, 'enable_sp':True, 'enable_sl':True, 'is_bond':False},
    {'code':'159915','mkt':'SZ','name':'创业板ETF',  'target':0.0625, 'enable_sp':True, 'enable_sl':True, 'is_bond':False},
    {'code':'512890','mkt':'SH','name':'红利低波ETF','target':0.0625, 'enable_sp':True, 'enable_sl':True, 'is_bond':False},
    {'code':'511380','mkt':'SH','name':'可转债ETF',  'target':0.125,  'enable_sp':True, 'enable_sl':True, 'is_bond':False},
    {'code':'511260','mkt':'SH','name':'10年国债ETF','target':0.125,  'enable_sp':False,'enable_sl':False,'is_bond':True},
    {'code':'511010','mkt':'SH','name':'5年国债ETF', 'target':0.125,  'enable_sp':False,'enable_sl':False,'is_bond':True},
    {'code':'511360','mkt':'SH','name':'短融ETF',    'target':0.0625, 'enable_sp':False,'enable_sl':False,'is_bond':True},
    {'code':'518880','mkt':'SH','name':'黄金ETF',    'target':0.1875, 'enable_sp':True, 'enable_sl':True, 'is_bond':False},
]

def round_lot(q):
    return int(max(0, math.floor(q / 100)) * 100)

# ========== GA 参数 ==========
POP_SIZE = 60       # 种群大小
GENERATIONS = 50    # 代数
TOURNAMENT_SIZE = 3
CROSSOVER_PROB = 0.9
MUTATION_PROB = 0.3  # 初始突变率
ELITE_COUNT = 2

# 染色体边界: [sp_trigger, sp_sell_pct, sl_trigger, observe_days, rb_threshold]
BOUNDS = [
    (0.05, 0.50),    # sp_trigger: 止盈触发阈值 5%-50%
    (0.30, 1.00),    # sp_sell_pct: 止盈卖出比例 30%-100%
    (0.10, 0.30),    # sl_trigger: 止损阈值 10%-30%（绝对值）
    (10, 60),        # observe_days: 观察天数
    (0.10, 0.40),    # rb_threshold: 再平衡偏离阈值 10%-40%
]

# 固定参数（不优化）
RB_FREQ = 'annual'  # 再平衡频率固定为年度（与 V3 一致）

# ========== 获取行情数据 ==========
print("=" * 60)
print("  达尔文 GA 优化器 — 遗传算法参数优化")
print("=" * 60)
print(f"  种群: {POP_SIZE} × 代: {GENERATIONS} = {POP_SIZE * (GENERATIONS + 1)} 次回测")
print(f"  参数: sp_trigger, sp_sell_pct, sl_trigger, observe_days, rb_threshold")
print(f"  适应度: CAGR - 0.5 × |max_drawdown|")
print()

print("获取行情数据 (2021-2025)...")
all_data = {}
for e in ETFS:
    ts_code = f"{e['code']}.{e['mkt']}"
    df = pro.fund_daily(ts_code=ts_code, start_date='20210101', end_date='20251231',
                        fields='trade_date,close,pre_close')
    if df is not None and len(df):
        df['trade_date'] = pd.to_datetime(df['trade_date'])
        df = df.sort_values('trade_date').reset_index(drop=True)
        all_data[e['code']] = df

all_dates = sorted(set().union(*[set(df['trade_date'].dt.date) for df in all_data.values()]))
print(f"  共 {len(all_dates)} 个交易日 ({all_dates[0]} ~ {all_dates[-1]})")

# 价格缓存（全局，所有回测共享）
price_cache = {}
for d in all_dates:
    price_cache[d] = {}
    for e in ETFS:
        df = all_data[e['code']]
        mask = df['trade_date'].dt.date == d
        price_cache[d][e['code']] = float(df.loc[mask, 'close'].iloc[0]) if mask.any() else None

# ========== 回测函数（参数化版） ==========
EVAL_COUNT = 0

def run_backtest(params):
    """params: [sp_trigger, sp_sell_pct, sl_trigger, observe_days, rb_threshold]"""
    global EVAL_COUNT
    EVAL_COUNT += 1

    sp_trigger, sp_sell_pct, sl_trigger, observe_days, rb_threshold = params
    sl_trigger = -abs(sl_trigger)  # 止损为负值
    observe_days = int(round(observe_days))

    cash = float(TOTAL)
    positions = {}
    for e in ETFS:
        p = {'code': e['code'], 'name': e['name'], 'target': e['target'],
             'enable_sp': e['enable_sp'], 'enable_sl': e['enable_sl'],
             'is_bond': e['is_bond'],
             'qty': 0, 'cost': 0.0, 'cost_total': 0.0,
             'price': 0.0, 'value': 0.0, 'pnl_pct': 0.0,
             'sp_triggered': False, 'sl_triggered': False,
             'cleared_date': None, 'observe_needed': 0}
        positions[e['code']] = p

    # 建仓
    first_date = all_dates[0]
    for e in ETFS:
        p = positions[e['code']]
        price = price_cache[first_date][e['code']]
        if price and price > 0:
            amt = TOTAL * e['target']
            qty = round_lot(amt / price)
            if qty > 0:
                spent = qty * price
                p['qty'] = qty
                p['cost'] = price
                p['cost_total'] = spent
                cash -= spent

    # 剩余现金买入短融
    p360 = positions['511360']
    price360 = price_cache[first_date]['511360']
    if price360 and price360 > 0 and cash > 1000:
        qty360 = round_lot(cash / price360)
        if qty360 > 0:
            spent360 = qty360 * price360
            p360['qty'] += qty360
            p360['cost_total'] += spent360
            p360['cost'] = p360['cost_total'] / p360['qty'] if p360['qty'] > 0 else 0
            cash -= spent360

    daily_values = []
    trade_count = 0

    def is_rb_date(date, idx, dates_list):
        if RB_FREQ is None:
            return False
        m = date.month
        if RB_FREQ == 'annual':
            if m != 12:
                return False
        elif RB_FREQ == 'semi':
            if m not in (6, 12):
                return False
        elif RB_FREQ == 'quarter':
            if m not in (3, 6, 9, 12):
                return False
        if idx + 1 < len(dates_list):
            if dates_list[idx + 1].month == m:
                return False
        return True

    for idx, date in enumerate(all_dates):
        # 更新价格
        for e in ETFS:
            p = positions[e['code']]
            pr = price_cache[date][e['code']]
            if pr and pr > 0:
                p['price'] = pr
                if p['qty'] > 0:
                    p['value'] = p['qty'] * pr
                    p['pnl_pct'] = (p['value'] - p['cost_total']) / p['cost_total'] if p['cost_total'] > 0 else 0
                else:
                    p['value'] = 0.0

        total_val = cash + sum(positions[e['code']]['value'] for e in ETFS)

        # 权重
        for e in ETFS:
            p = positions[e['code']]
            p['weight'] = p['value'] / total_val if total_val > 0 else 0

        # ===== 止盈检查 =====
        if sp_trigger is not None:
            for e in ETFS:
                p = positions[e['code']]
                if not p['enable_sp'] or p['qty'] <= 0 or p['sp_triggered']:
                    continue
                if p['pnl_pct'] >= sp_trigger:
                    sell_qty = round_lot(p['qty'] * sp_sell_pct)
                    if sell_qty > 0:
                        sold_amt = sell_qty * p['price']
                        cost_part = p['cost'] * sell_qty
                        p['qty'] -= sell_qty
                        p['cost_total'] -= cost_part
                        p['cost'] = p['cost_total'] / p['qty'] if p['qty'] > 0 else 0
                        cash += sold_amt
                        p['sp_triggered'] = True
                        trade_count += 1
                        # 止盈资金转入短融
                        p360_2 = positions['511360']
                        pr360 = price_cache[date]['511360']
                        if pr360 and pr360 > 0 and sold_amt > 1000:
                            q360 = round_lot(sold_amt / pr360)
                            if q360 > 0:
                                sp360 = q360 * pr360
                                p360_2['qty'] += q360
                                p360_2['cost_total'] += sp360
                                p360_2['cost'] = p360_2['cost_total'] / p360_2['qty'] if p360_2['qty'] > 0 else 0
                                cash -= sp360

        # ===== 止损检查 =====
        if sl_trigger is not None:
            for e in ETFS:
                p = positions[e['code']]
                if not p['enable_sl'] or p['qty'] <= 0 or p['sl_triggered']:
                    continue
                if p['pnl_pct'] <= sl_trigger:
                    sold_amt = p['qty'] * p['price']
                    cash += sold_amt
                    p['sl_triggered'] = True
                    p['cleared_date'] = date
                    p['observe_needed'] = observe_days   # GA 优化的观察天数
                    p['qty'] = 0
                    p['cost'] = 0
                    p['cost_total'] = 0
                    trade_count += 1

        # ===== 观察期满重新建仓 =====
        for e in ETFS:
            p = positions[e['code']]
            if p['qty'] > 0 or not p['sl_triggered'] or p['cleared_date'] is None:
                continue
            days_since = (date - p['cleared_date']).days
            est_td = int(days_since * 0.7)
            if est_td >= p['observe_needed']:
                target_amt = TOTAL * p['target']
                pr = price_cache[date][p['code']]
                if pr and pr > 0 and target_amt > 1000:
                    if cash < target_amt:
                        p360_3 = positions['511360']
                        pr360 = price_cache[date]['511360']
                        if pr360 and pr360 > 0 and p360_3['qty'] > 0:
                            need = target_amt - cash
                            sell_q = round_lot(need / pr360)
                            if sell_q > 0:
                                sa = sell_q * pr360
                                cp = p360_3['cost'] * sell_q
                                p360_3['qty'] -= sell_q
                                p360_3['cost_total'] -= cp
                                p360_3['cost'] = p360_3['cost_total'] / p360_3['qty'] if p360_3['qty'] > 0 else 0
                                cash += sa
                    buy_amt = min(target_amt, cash)
                    qty_buy = round_lot(buy_amt / pr)
                    if qty_buy > 0:
                        spent = qty_buy * pr
                        p['qty'] = qty_buy
                        p['cost'] = pr
                        p['cost_total'] = spent
                        p['sl_triggered'] = False
                        p['sp_triggered'] = False
                        p['cleared_date'] = None
                        p['observe_needed'] = 0
                        cash -= spent
                        trade_count += 1

        # ===== 再平衡 =====
        if is_rb_date(date, idx, all_dates):
            triggers = False
            deviations = []
            for e in ETFS:
                p = positions[e['code']]
                if p['target'] > 0:
                    act_w = p['value'] / total_val if total_val > 0 else 0
                    rel_dev = (act_w - p['target']) / p['target']
                    deviations.append((e['code'], e['name'], act_w, p['target'], rel_dev, p))
                    if abs(rel_dev) > rb_threshold:
                        triggers = True

            if triggers:
                current_total = total_val
                # 卖超配
                for code, name, act_w, tgt_w, rel_dev, p in deviations:
                    if rel_dev > 0.01 and p['qty'] > 0:
                        excess = p['value'] - current_total * tgt_w
                        if excess > 100:
                            pr = price_cache[date][code]
                            if pr and pr > 0:
                                sq = round_lot(excess / pr)
                                if sq > 0:
                                    sa = sq * pr
                                    cp = p['cost'] * sq
                                    p['qty'] -= sq
                                    p['cost_total'] -= cp
                                    p['cost'] = p['cost_total'] / p['qty'] if p['qty'] > 0 else 0
                                    cash += sa
                                    trade_count += 1

                total_val = cash + sum(positions[e['code']]['value'] for e in ETFS)

                # 买低配
                for code, name, act_w, tgt_w, rel_dev, p in sorted(deviations, key=lambda x: x[4]):
                    if rel_dev < -0.01 and p['qty'] >= 0:
                        target_amount = total_val * tgt_w
                        shortfall = target_amount - p['value']
                        if shortfall > 100:
                            pr = price_cache[date][code]
                            if pr and pr > 0:
                                if cash < shortfall:
                                    p360_4 = positions['511360']
                                    pr360 = price_cache[date]['511360']
                                    if pr360 and pr360 > 0 and p360_4['qty'] > 0:
                                        need = shortfall - cash
                                        sq4 = round_lot(need / pr360)
                                        if sq4 > 0:
                                            sa4 = sq4 * pr360
                                            cp4 = p360_4['cost'] * sq4
                                            p360_4['qty'] -= sq4
                                            p360_4['cost_total'] -= cp4
                                            p360_4['cost'] = p360_4['cost_total'] / p360_4['qty'] if p360_4['qty'] > 0 else 0
                                            cash += sa4
                                buy_amt = min(shortfall, cash)
                                qb = round_lot(buy_amt / pr)
                                if qb > 0:
                                    spent = qb * pr
                                    p['qty'] += qb
                                    p['cost_total'] += spent
                                    p['cost'] = p['cost_total'] / p['qty'] if p['qty'] > 0 else 0
                                    cash -= spent
                                    trade_count += 1

                # 剩余现金入短融
                p360_5 = positions['511360']
                pr360 = price_cache[date]['511360']
                if cash > 5000 and pr360 and pr360 > 0:
                    qb5 = round_lot(cash / pr360)
                    if qb5 > 0:
                        spent5 = qb5 * pr360
                        p360_5['qty'] += qb5
                        p360_5['cost_total'] += spent5
                        p360_5['cost'] = p360_5['cost_total'] / p360_5['qty'] if p360_5['qty'] > 0 else 0
                        cash -= spent5

        daily_values.append(total_val)

    # 计算指标
    start_val = float(TOTAL)
    end_val = daily_values[-1] if daily_values else start_val
    years = (all_dates[-1] - all_dates[0]).days / 365.25
    cagr = ((end_val / start_val) ** (1 / years) - 1) * 100 if years > 0 else 0
    total_ret = (end_val / start_val - 1) * 100

    peak_val = 0
    max_dd = 0
    for v in daily_values:
        if v > peak_val:
            peak_val = v
        dd = (v - peak_val) / peak_val
        if dd < max_dd:
            max_dd = dd
    max_dd_pct = max_dd * 100

    # 计算 Sharpe-like（日收益率年化 / 年化波动率，无风险利率=0）
    daily_returns = []
    for i in range(1, len(daily_values)):
        if daily_values[i-1] > 0:
            daily_returns.append(daily_values[i] / daily_values[i-1] - 1)
    if len(daily_returns) > 0 and np.std(daily_returns) > 0:
        sharpe = np.mean(daily_returns) / np.std(daily_returns) * np.sqrt(252)
    else:
        sharpe = 0

    # 适应度：CAGR 减去回撤惩罚
    fitness = cagr - 0.5 * abs(max_dd_pct)

    return {
        'cagr': round(cagr, 4),
        'total_ret': round(total_ret, 4),
        'max_dd': round(max_dd_pct, 4),
        'sharpe': round(sharpe, 4),
        'trades': trade_count,
        'fitness': round(fitness, 4),
        'final_val': round(end_val, 2),
    }


# ========== 遗传算法核心 ==========

def clamp(val, lo, hi):
    return max(lo, min(hi, val))

def random_individual():
    """生成随机个体"""
    return [random.uniform(lo, hi) for lo, hi in BOUNDS]

def tournament_select(population, fitnesses):
    """锦标赛选择"""
    best_idx = None
    best_fit = -float('inf')
    for _ in range(TOURNAMENT_SIZE):
        idx = random.randrange(len(population))
        if fitnesses[idx] > best_fit:
            best_fit = fitnesses[idx]
            best_idx = idx
    return population[best_idx][:]

def sbx_crossover(p1, p2):
    """模拟二进制交叉 (Simulated Binary Crossover)"""
    if random.random() > CROSSOVER_PROB:
        return p1[:], p2[:]
    eta = 15  # 分布指数
    c1, c2 = p1[:], p2[:]
    for i in range(len(BOUNDS)):
        if random.random() < 0.5:
            lo, hi = BOUNDS[i]
            if abs(p2[i] - p1[i]) > 1e-10:
                if p1[i] < p2[i]:
                    y1, y2 = p1[i], p2[i]
                else:
                    y1, y2 = p2[i], p1[i]
                # 计算 beta
                beta = 1.0 + 2.0 * (y1 - lo) / (y2 - y1) if (y2 - y1) > 0 else 1.0
                alpha = 2.0 - beta ** -(eta + 1)
                u = random.random()
                if u <= 1.0 / alpha:
                    beta_q = (u * alpha) ** (1.0 / (eta + 1))
                else:
                    beta_q = (1.0 / (2.0 - u * alpha)) ** (1.0 / (eta + 1))
                child1 = 0.5 * ((y1 + y2) - beta_q * (y2 - y1))
                child2 = 0.5 * ((y1 + y2) + beta_q * (y2 - y1))
                c1[i] = clamp(child1, lo, hi)
                c2[i] = clamp(child2, lo, hi)
    return c1, c2

def polynomial_mutation(ind, gen):
    """多项式突变，突变强度随代数衰减"""
    # 衰减突变率
    decay = 1.0 - gen / GENERATIONS
    current_mut_prob = MUTATION_PROB * (0.4 + 0.6 * decay)
    eta_m = 20  # 突变分布指数
    for i in range(len(BOUNDS)):
        if random.random() < current_mut_prob:
            lo, hi = BOUNDS[i]
            delta = (hi - lo) * 0.1  # 初始 delta
            u = random.random()
            if u < 0.5:
                delta_q = (2 * u) ** (1.0 / (eta_m + 1)) - 1.0
            else:
                delta_q = 1.0 - (2 * (1 - u)) ** (1.0 / (eta_m + 1))
            ind[i] = clamp(ind[i] + delta_q * delta, lo, hi)


# ========== 主优化循环 ==========
print("\n" + "=" * 60)
print("  开始遗传算法优化")
print("=" * 60)

# 初始化种群
random.seed(42)
np.random.seed(42)

population = [random_individual() for _ in range(POP_SIZE)]
convergence = []  # 每代最优适应度

print(f"\n{'代':>4}  {'最优CAGR':>8}  {'最优回撤':>8}  {'最优适应度':>10}  {'平均适应度':>10}  {'耗时':>8}")
print("-" * 60)

gen_start_time = time.time()

for gen in range(GENERATIONS + 1):  # +1 包含初始种群评估
    gen_t0 = time.time()

    # 评估当前种群
    gen_results = []
    for ind in population:
        result = run_backtest(ind)
        gen_results.append(result)

    fitnesses = [r['fitness'] for r in gen_results]

    # 找最优
    best_idx = max(range(len(fitnesses)), key=lambda i: fitnesses[i])
    best_ind = population[best_idx][:]
    best_result = gen_results[best_idx]
    avg_fitness = sum(fitnesses) / len(fitnesses)

    convergence.append({
        'generation': gen,
        'best_fitness': best_result['fitness'],
        'best_cagr': best_result['cagr'],
        'best_max_dd': best_result['max_dd'],
        'avg_fitness': round(avg_fitness, 4),
    })

    gen_elapsed = time.time() - gen_t0
    total_elapsed = time.time() - gen_start_time

    print(f"{gen:>4}  {best_result['cagr']:>8.2f}%  {best_result['max_dd']:>8.2f}%  "
          f"{best_result['fitness']:>10.4f}  {avg_fitness:>10.4f}  {total_elapsed:>7.1f}s")

    if gen >= GENERATIONS:
        break

    # ==== 生成下一代 ====
    # 精英保留
    sorted_idx = sorted(range(len(fitnesses)), key=lambda i: fitnesses[i], reverse=True)
    new_population = [population[sorted_idx[i]][:] for i in range(ELITE_COUNT)]

    # 填充剩余个体
    while len(new_population) < POP_SIZE:
        # 选择
        parent1 = tournament_select(population, fitnesses)
        parent2 = tournament_select(population, fitnesses)

        # 交叉
        child1, child2 = sbx_crossover(parent1, parent2)

        # 突变
        polynomial_mutation(child1, gen)
        polynomial_mutation(child2, gen)

        new_population.append(child1)
        if len(new_population) < POP_SIZE:
            new_population.append(child2)

    population = new_population

total_time = time.time() - gen_start_time
print(f"\n  总耗时: {total_time:.1f}s ({total_time/60:.1f}min)")
print(f"  总评估: {EVAL_COUNT} 次回测")

# ========== 最终结果 ==========
print("\n" + "=" * 60)
print("  GA 优化结果")
print("=" * 60)

sp_trigger, sp_sell_pct, sl_trigger, observe_days, rb_threshold = best_ind

print(f"""
  最优参数:
    止盈触发:  +{sp_trigger*100:.1f}%  卖出 {sp_sell_pct*100:.0f}%
    止损触发:  -{abs(sl_trigger)*100:.1f}%
    观察天数:  {int(round(observe_days))} 个交易日
    再平衡:    年度 ±{rb_threshold*100:.0f}%

  回测指标 (2021-2025, 400万):
    CAGR:      {best_result['cagr']:.2f}%
    总收益:    {best_result['total_ret']:.2f}%
    最大回撤:  {best_result['max_dd']:.2f}%
    Sharpe:    {best_result['sharpe']:.4f}
    交易笔数:  {best_result['trades']}
    终值:      ¥{best_result['final_val']:,.0f}
""")

# ========== 对比基准 ==========
print("=" * 60)
print("  四方案对比")
print("=" * 60)

# 买入持有基准
buyhold_cagr = 5.81
buyhold_ret = 32.64
buyhold_dd = -12.43

# 网格最优
grid_cagr = 10.68
grid_ret = 65.88
grid_dd = -12.88

# V3
v3_cagr = 5.39
v3_ret = 28.72
v3_dd = -10.32

print(f"""
  {'':20} {'CAGR':>8} {'总收益':>8} {'最大回撤':>8} {'Sharpe':>8}
  {'-'*52}
  {'GA最优':20} {best_result['cagr']:>7.2f}% {best_result['total_ret']:>7.2f}% {best_result['max_dd']:>7.2f}% {best_result['sharpe']:>7.4f}
  {'网格最优(sp30/sl15)':20} {grid_cagr:>7.2f}% {grid_ret:>7.2f}% {grid_dd:>7.2f}% {'—':>8}
  {'V3(sp20/sl15)':20} {v3_cagr:>7.2f}% {v3_ret:>7.2f}% {v3_dd:>7.2f}% {'—':>8}
  {'买入持有':20} {buyhold_cagr:>7.2f}% {buyhold_ret:>7.2f}% {buyhold_dd:>7.2f}% {'—':>8}
""")

# ========== 保存结果 ==========
output = {
    'algorithm': 'Genetic Algorithm (hand-rolled)',
    'description': 'Tournament selection + SBX crossover + polynomial mutation + elitism',
    'config': {
        'population_size': POP_SIZE,
        'generations': GENERATIONS,
        'total_evaluations': EVAL_COUNT,
        'tournament_size': TOURNAMENT_SIZE,
        'crossover_prob': CROSSOVER_PROB,
        'mutation_prob': MUTATION_PROB,
        'elite_count': ELITE_COUNT,
        'bounds': {k: list(v) for k, v in zip(
            ['sp_trigger', 'sp_sell_pct', 'sl_trigger', 'observe_days', 'rb_threshold'], BOUNDS)},
        'data_period': f'{all_dates[0]} ~ {all_dates[-1]}',
        'trading_days': len(all_dates),
        'rb_freq': RB_FREQ,
    },
    'best_params': {
        'sp_trigger': round(sp_trigger, 4),
        'sp_sell_pct': round(sp_sell_pct, 4),
        'sl_trigger': round(abs(sl_trigger), 4),
        'observe_days': int(round(observe_days)),
        'rb_threshold': round(rb_threshold, 4),
    },
    'best_metrics': best_result,
    'convergence': convergence,
    'comparison': {
        'ga_best': {'cagr': best_result['cagr'], 'total_ret': best_result['total_ret'],
                     'max_dd': best_result['max_dd'], 'sharpe': best_result['sharpe']},
        'grid_best': {'cagr': grid_cagr, 'total_ret': grid_ret, 'max_dd': grid_dd,
                       'params': 'sp30/sl15/semi-annual-10'},
        'v3_live': {'cagr': v3_cagr, 'total_ret': v3_ret, 'max_dd': v3_dd,
                      'params': 'sp20/sl15/annual-20/observe20'},
        'buy_hold': {'cagr': buyhold_cagr, 'total_ret': buyhold_ret, 'max_dd': buyhold_dd},
    },
}

out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'ga_optimization_results.json')
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)
print(f"\n结果已保存: {out_path}")

# ========== 更新 Excel GA优化方案 sheet ==========
print("\n更新 Excel GA优化方案 sheet...")
try:
    from openpyxl import load_workbook
    xlsx_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                             '400万资产配置组合2026.xlsx')
    wb = load_workbook(xlsx_path)
    ws = wb['GA优化方案']

    sp_label = f">= +{sp_trigger*100:.1f}%，卖出{sp_sell_pct*100:.0f}%"
    sl_label = f"<= -{abs(sl_trigger)*100:.1f}%"
    ob_label = f"{int(round(observe_days))} 个交易日"

    # 更新参数区域（根据实际 Excel 结构定位）
    # Row mapping: B5=止盈, B6=止损, B7=观察期, B8=再平衡频率, B9=阀值
    ws['B5'] = sp_label
    ws['B6'] = sl_label
    ws['B7'] = ob_label
    ws['B9'] = f"+/-{rb_threshold*100:.0f}%"

    # 更新回测指标（B13-B18）
    ws['B13'] = f"+{best_result['cagr']:.2f}%"
    ws['B14'] = f"+{best_result['total_ret']:.2f}%"
    ws['B15'] = f"{best_result['max_dd']:.2f}%"
    ws['B16'] = f"{best_result['sharpe']:.4f}"
    ws['B17'] = f"¥{best_result['final_val']:,.0f}"
    ws['B18'] = best_result['trades']

    wb.save(xlsx_path)
    print("  已更新 GA优化方案 sheet")
except Exception as ex:
    print(f"  更新 Excel 失败: {ex}")
    print("  JSON 结果已保存，可手动更新")

print("\n" + "=" * 60)
print("  达尔文 GA 优化完成!")
print("=" * 60)
