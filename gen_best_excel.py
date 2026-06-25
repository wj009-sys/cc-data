#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""生成最优组合回测报表"""
import json, sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open('backtest_rules_result.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

trades = data['trades']
annual = data['annual_returns']

wb = Workbook()

hfont = Font(name='Microsoft YaHei', bold=True, size=11, color='FFFFFF')
hfill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
halign = Alignment(horizontal='center', vertical='center')
bfont = Font(name='Microsoft YaHei', size=10)
bbfont = Font(name='Microsoft YaHei', bold=True, size=10)
gfont = Font(name='Microsoft YaHei', size=10, color='006100')
gbfont = Font(name='Microsoft YaHei', bold=True, size=11, color='006100')
rfont = Font(name='Microsoft YaHei', size=10, color='9C0006')
ca = Alignment(horizontal='center', vertical='center')
la = Alignment(horizontal='left', vertical='center')
ra = Alignment(horizontal='right', vertical='center')
bdr = Border(left=Side(style='thin',color='D0D0D0'), right=Side(style='thin',color='D0D0D0'),
             top=Side(style='thin',color='D0D0D0'), bottom=Side(style='thin',color='D0D0D0'))
buy_f = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
sell_f = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
reb_f = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')

# ========================================
# Sheet 1: 交易流水
# ========================================
ws1 = wb.active
ws1.title = '交易流水'
headers = ['日期', '代码', '名称', '操作', '单价', '数量', '成交金额', '原因']
for col, t in enumerate(headers, 1):
    c = ws1.cell(row=1, column=col, value=t)
    c.font, c.fill, c.alignment, c.border = hfont, hfill, halign, bdr

ts = sorted(trades, key=lambda x: (x['date'], x['code']))
for i, t in enumerate(ts):
    r = i + 2
    if '再平衡' in t['action']:
        fill = reb_f
    elif '买入' in t['action']:
        fill = buy_f
    else:
        fill = sell_f
    vals = [t['date'], t['code'], t['name'], t['action'],
            t['price'], t['qty'], t['amount'], t['reason']]
    for col, v in enumerate(vals, 1):
        c = ws1.cell(row=r, column=col, value=v)
        c.font, c.border, c.fill = bfont, bdr, fill
        if col in (3, 8):
            c.alignment = la
        elif col in (5, 7):
            c.alignment = ra
        else:
            c.alignment = ca
        if col == 5:
            c.number_format = '0.0000'
        elif col == 7:
            c.number_format = '#,##0.00'
        elif col == 6:
            c.number_format = '#,##0'

for i, w in enumerate([12, 8, 14, 14, 10, 10, 14, 50], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.freeze_panes = 'A2'
ws1.auto_filter.ref = f'A1:H{len(ts)+1}'

# ========================================
# Sheet 2: 年度统计
# ========================================
ws2 = wb.create_sheet('年度统计')
trades_by_year = {}
for t in ts:
    y = t['date'][:4]
    trades_by_year[y] = trades_by_year.get(y, 0) + 1

for col, t in enumerate(['年份', '年初市值', '年末市值', '年收益率', '最大回撤', '年化波动', '交易笔数'], 1):
    c = ws2.cell(row=1, column=col, value=t)
    c.font, c.fill, c.alignment, c.border = hfont, hfill, halign, bdr

for i, ar in enumerate(annual):
    r = i + 2
    y = ar['year']
    vals = [y, ar['start_value'], ar['end_value'],
            ar['return_pct'] / 100, ar['max_drawdown'] / 100,
            ar['volatility_pct'] / 100, trades_by_year.get(str(y), 0)]
    for col, v in enumerate(vals, 1):
        c = ws2.cell(row=r, column=col, value=v)
        c.font, c.border, c.alignment = bfont, bdr, ca
        if col == 1:
            c.number_format = '0'
        elif col in (2, 3):
            c.number_format = '#,##0'
        elif col in (4, 5, 6):
            c.number_format = '0.00%'
        elif col == 7:
            c.number_format = '0'
        if col == 4:
            c.font = gfont if v >= 0 else rfont

tr = len(annual) + 2
c1 = ws2.cell(row=tr, column=1, value='总计')
c1.font, c1.border, c1.alignment = bbfont, bdr, ca
for col, v in enumerate([4000000, annual[-1]['end_value']], 2):
    c = ws2.cell(row=tr, column=col, value=v)
    c.font, c.border, c.alignment = bbfont, bdr, ca
    c.number_format = '#,##0'
cc = ws2.cell(row=tr, column=4, value=data['summary']['cagr_pct'] / 100)
cc.font, cc.border, cc.alignment = gbfont, bdr, ca
cc.number_format = '0.00%'
tc = ws2.cell(row=tr, column=5, value=data['summary']['total_return_pct'] / 100)
tc.font, tc.border, tc.alignment = bbfont, bdr, ca
tc.number_format = '0.00%'
c6 = ws2.cell(row=tr, column=6, value='CAGR')
c6.font, c6.border, c6.alignment = bbfont, bdr, ca
c7 = ws2.cell(row=tr, column=7, value=len(ts))
c7.font, c7.border, c7.alignment = bbfont, bdr, ca
for i, w in enumerate([8, 14, 14, 10, 10, 10, 10], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ========================================
# Sheet 3: 四方案对比
# ========================================
ws3 = wb.create_sheet('四方案对比')
h3 = ['年份', '买入持有', '原纪律(阶梯止盈)', '用户方案(宽)', '最优(推荐)']
for col, t in enumerate(h3, 1):
    c = ws3.cell(row=1, column=col, value=t)
    c.font, c.fill, c.alignment, c.border = hfont, hfill, halign, bdr

old_results = {2021: -2.74, 2022: -2.42, 2023: -1.42, 2024: 11.05, 2025: 23.07}
user_results = {2021: 0.97, 2022: -6.82, 2023: 0.46, 2024: 15.93, 2025: 22.05}
best_results = {ar['year']: ar['return_pct'] for ar in annual}
bh = {2021: -1.33, 2022: -5.77, 2023: 0.63, 2024: 14.31, 2025: 23.16}

for i, y in enumerate([2021, 2022, 2023, 2024, 2025], 2):
    c = ws3.cell(row=i, column=1, value=y)
    c.font, c.border, c.alignment = bfont, bdr, ca
    for col, v in enumerate([bh[y], old_results[y], user_results[y], best_results[y]], 2):
        cell = ws3.cell(row=i, column=col, value=v / 100)
        cell.font, cell.border, cell.alignment = bfont, bdr, ca
        cell.number_format = '0.00%'
        cell.font = gfont if v >= 0 else rfont

cr = 7
c1 = ws3.cell(row=cr, column=1, value='CAGR')
c1.font, c1.border, c1.alignment = bbfont, bdr, ca
cagr_vals = [5.81, 4.79, 5.74, data['summary']['cagr_pct']]
for col, v in enumerate(cagr_vals, 2):
    cell = ws3.cell(row=cr, column=col, value=v / 100)
    cell.font, cell.border, cell.alignment = gbfont, bdr, ca
    cell.number_format = '0.00%'

tr4 = 8
c1 = ws3.cell(row=tr4, column=1, value='交易笔数')
c1.font, c1.border, c1.alignment = bbfont, bdr, ca
for col, v in enumerate([0, 171, 80, 90], 2):
    cell = ws3.cell(row=tr4, column=col, value=v)
    cell.font, cell.border, cell.alignment = bbfont, bdr, ca

for i, w in enumerate([10, 14, 14, 14, 14], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ========================================
# 保存
# ========================================
fp = 'backtest_best_trades.xlsx'
wb.save(fp)
print(f'OK: {fp}')
