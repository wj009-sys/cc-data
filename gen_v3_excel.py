#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""生成v3最优组合报表"""
import json, sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open('backtest_best_result.json','r',encoding='utf-8') as f:
    data = json.load(f)
trades = data['trades']; annual = data['annual_returns']

wb = Workbook()

hfont=Font(name='Microsoft YaHei',bold=True,size=11,color='FFFFFF')
hfill=PatternFill(start_color='1F4E79',end_color='1F4E79',fill_type='solid')
halign=Alignment(horizontal='center',vertical='center')
bfont=Font(name='Microsoft YaHei',size=10)
bbfont=Font(name='Microsoft YaHei',bold=True,size=10)
gfont=Font(name='Microsoft YaHei',size=10,color='006100')
rfont=Font(name='Microsoft YaHei',size=10,color='9C0006')
gbfont=Font(name='Microsoft YaHei',bold=True,size=11,color='006100')
ca,la,ra=Alignment(horizontal='center',vertical='center'),Alignment(horizontal='left',vertical='center'),Alignment(horizontal='right',vertical='center')
bdr=Border(left=Side(style='thin',color='D0D0D0'),right=Side(style='thin',color='D0D0D0'),top=Side(style='thin',color='D0D0D0'),bottom=Side(style='thin',color='D0D0D0'))
buy_f=PatternFill(start_color='E8F5E9',end_color='E8F5E9',fill_type='solid')
sell_f=PatternFill(start_color='FFEBEE',end_color='FFEBEE',fill_type='solid')
reb_f=PatternFill(start_color='E3F2FD',end_color='E3F2FD',fill_type='solid')

# Sheet 1: 交易流水
ws1=wb.active; ws1.title='交易流水'
for c,t in enumerate(['日期','代码','名称','操作','单价','数量','成交金额','原因'],1):
    cell=ws1.cell(row=1,column=c,value=t); cell.font,cell.fill,cell.alignment,cell.border=hfont,hfill,halign,bdr
ts=sorted(trades,key=lambda x:(x['date'],x['code']))
for i,t in enumerate(ts):
    r=i+2
    fill=reb_f if '再平衡' in t['action'] else(buy_f if '买入' in t['action'] else sell_f)
    for c,v in enumerate([t['date'],t['code'],t['name'],t['action'],t['price'],t['qty'],t['amount'],t['reason']],1):
        cell=ws1.cell(row=r,column=c,value=v); cell.font,cell.border,cell.fill=bfont,bdr,fill
        cell.alignment=la if c in(3,8)else(ra if c in(5,7)else ca)
        if c==5: cell.number_format='0.0000'
        elif c==7: cell.number_format='#,##0'
        elif c==6: cell.number_format='#,##0'
for i,w in enumerate([12,8,12,8,8,8,10,45],1): ws1.column_dimensions[get_column_letter(i)].width=w
ws1.freeze_panes='A2'; ws1.auto_filter.ref=f'A1:H{len(ts)+1}'

# Sheet 2: 年度统计
ws2=wb.create_sheet('年度统计')
ty={}
for t in ts:
    y=t['date'][:4]; ty[y]=ty.get(y,0)+1
for c,t in enumerate(['年份','年初市值','年末市值','年收益率','交易笔数'],1):
    cell=ws2.cell(row=1,column=c,value=t); cell.font,cell.fill,cell.alignment,cell.border=hfont,hfill,halign,bdr
for i,ar in enumerate(annual):
    r=i+2; y=ar['year']
    for c,v in enumerate([y,ar['start'],ar['end'],ar['return']/100,ty.get(str(y),0)],1):
        cell=ws2.cell(row=r,column=c,value=v); cell.font,cell.border,cell.alignment=bfont,bdr,ca
        if c==1: cell.number_format='0'
        elif c in(2,3): cell.number_format='#,##0'
        elif c==4: cell.number_format='0.00%'; cell.font=gfont if v>=0 else rfont
        elif c==5: cell.number_format='0'
tr=len(annual)+2
for c,v in enumerate(['总计',4000000,annual[-1]['end'],data['summary']['cagr_pct']/100,len(ts)],1):
    cell=ws2.cell(row=tr,column=c,value=v); cell.font,cell.border,cell.alignment=bbfont,bdr,ca
    if c==1: pass
    elif c in(2,3): cell.number_format='#,##0'
    elif c==4: cell.number_format='0.00%'; cell.font=gbfont
    elif c==5: cell.number_format='0'
for i,w in enumerate([8,12,12,10,10],1): ws2.column_dimensions[get_column_letter(i)].width=w

# Sheet 3: 多方案对比
ws3=wb.create_sheet('多方案对比')
h3=['年份','买入持有','原纪律','用户方案','最优v2(观察40日)','最优v3(观察20日+纯现金)']
for c,t in enumerate(h3,1):
    cell=ws3.cell(row=1,column=c,value=t); cell.font,cell.fill,cell.alignment,cell.border=hfont,hfill,halign,bdr
bh={2021:-1.33,2022:-5.77,2023:0.63,2024:14.31,2025:23.16}
old={2021:-2.74,2022:-2.42,2023:-1.42,2024:11.05,2025:23.07}
user={2021:0.97,2022:-6.82,2023:0.46,2024:15.93,2025:22.05}
v2={2021:0.63,2022:-1.14,2023:1.11,2024:5.84,2025:21.68}
v3={ar['year']:ar['return'] for ar in annual}
for i,y in enumerate([2021,2022,2023,2024,2025],2):
    ws3.cell(row=i,column=1,value=y).font=bbfont; ws3.cell(row=i,column=1).border=bdr; ws3.cell(row=i,column=1).alignment=ca
    for c,v in enumerate([bh[y],old[y],user[y],v2[y],v3[y]],2):
        cell=ws3.cell(row=i,column=c,value=v/100)
        cell.font,cell.border,cell.alignment=bfont,bdr,ca; cell.number_format='0.00%'
        cell.font=gfont if v>=0 else rfont
cr=7
ws3.cell(row=cr,column=1,value='CAGR').font=bbfont; ws3.cell(row=cr,column=1).border=bdr; ws3.cell(row=cr,column=1).alignment=ca
for c,v in enumerate([5.81,4.79,5.74,5.09,data['summary']['cagr_pct']],2):
    cell=ws3.cell(row=cr,column=c,value=v/100); cell.font,cell.border,cell.alignment=gbfont,bdr,ca; cell.number_format='0.00%'
tr2=8
ws3.cell(row=tr2,column=1,value='交易笔数').font=bbfont; ws3.cell(row=tr2,column=1).border=bdr; ws3.cell(row=tr2,column=1).alignment=ca
for c,v in enumerate([0,161,80,90,69],2):
    cell=ws3.cell(row=tr2,column=c,value=v); cell.font,cell.border,cell.alignment=bbfont,bdr,ca
for i,w in enumerate([10,12,12,12,14,16],1): ws3.column_dimensions[get_column_letter(i)].width=w

fp='backtest_v3_report.xlsx'; wb.save(fp)
print(f'OK: {fp}')
