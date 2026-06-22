#!/usr/bin/env python
"""回补 2026-06-16 ~ 2026-06-17 的日报数据"""
import sys, io, os
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

import tushare as ts
from openpyxl import load_workbook
from datetime import date, timedelta
from pathlib import Path

FILE_PATH = Path(r'd:\cc-data\500万资产配置组合2026.xlsx')

HOLDINGS = [
    {"code": 510300, "name": "沪深300ETF",        "ts": "510300.SH", "high_vol": False, "type": "equity"},
    {"code": 588050, "name": "科创50ETF",         "ts": "588050.SH", "high_vol": True,  "type": "equity"},
    {"code": 159915, "name": "创业板ETF",         "ts": "159915.SZ", "high_vol": True,  "type": "equity"},
    {"code": 512890, "name": "红利低波ETF",       "ts": "512890.SH", "high_vol": False, "type": "equity"},
    {"code": 511380, "name": "可转债ETF",         "ts": "511380.SH", "high_vol": False, "type": "equity"},
    {"code": 511260, "name": "10年国债ETF",       "ts": "511260.SH", "high_vol": False, "type": "bond"},
    {"code": 511010, "name": "5年国债ETF",        "ts": "511010.SH", "high_vol": False, "type": "bond"},
    {"code": 511360, "name": "短融ETF",           "ts": "511360.SH", "high_vol": False, "type": "bond"},
    {"code": 518880, "name": "黄金ETF",           "ts": "518880.SH", "high_vol": False, "type": "commodity"},
]
HOLDING_MAP = {h["code"]: h for h in HOLDINGS}
CODE_ROWS = [2, 3, 4, 5, 6, 7, 8, 9, 10]

pro = ts.pro_api()

def fetch_prices_for_date(ts_codes, trade_date_str):
    """获取指定交易日所有品种价格"""
    result = {}
    for code in ts_codes:
        try:
            df = pro.fund_daily(ts_code=code, trade_date=trade_date_str)
            if not df.empty:
                row = df.iloc[0]
                result[code] = {
                    "close": float(row["close"]),
                    "pct_chg": float(row["pct_chg"]),
                }
            else:
                start = (__import__('datetime').datetime.strptime(trade_date_str, "%Y%m%d") - timedelta(days=5)).strftime("%Y%m%d")
                df2 = pro.fund_daily(ts_code=code, start_date=start, end_date=trade_date_str)
                if not df2.empty:
                    row = df2.iloc[0]
                    result[code] = {
                        "close": float(row["close"]),
                        "pct_chg": float(row["pct_chg"]),
                    }
        except Exception as e:
            print(f"  [WARN] {code}: {e}")
    return result


def add_daily_snapshot(wb, prices, trade_date_str):
    """为指定日期追加日报数据（不修改汇总表）"""
    ws_summary = wb["汇总"]
    ws_daily = wb["日报"]

    # 从汇总表读取持仓信息
    holdings_data = []
    total_market_value = 0
    total_cost_value = 0

    for i, h in enumerate(HOLDINGS):
        row = CODE_ROWS[i]
        ts_code = h["ts"]
        cost_price = ws_summary.cell(row=row, column=8).value
        quantity = ws_summary.cell(row=row, column=9).value or 0
        target_weight = ws_summary.cell(row=row, column=5).value or 0

        if ts_code in prices:
            close = prices[ts_code]["close"]
            pct_chg = prices[ts_code]["pct_chg"]
        else:
            close = ws_summary.cell(row=row, column=11).value or 0
            pct_chg = 0

        if quantity and quantity > 0 and cost_price and cost_price > 0:
            market_value = close * quantity
            cost_value = cost_price * quantity
            total_market_value += market_value
            total_cost_value += cost_value
            pnl = round((close - cost_price) * quantity, 2)
            pnl_pct = round((close - cost_price) / cost_price, 4)
        else:
            market_value = 0
            cost_value = 0
            pnl = 0
            pnl_pct = 0

        holdings_data.append({
            "code": h["code"], "name": h["name"], "close": close,
            "cost_price": cost_price, "quantity": quantity,
            "pct_chg": pct_chg, "pnl": pnl, "pnl_pct": pnl_pct,
            "market_value": market_value, "cost_value": cost_value,
            "target_weight": target_weight,
        })

    # 计算权重
    for hd in holdings_data:
        if total_market_value > 0 and hd["quantity"] and hd["quantity"] > 0:
            hd["weight"] = round(hd["market_value"] / total_market_value, 4)
        else:
            hd["weight"] = 0
        hd["deviation"] = round(hd["weight"] - hd["target_weight"], 4)

    # 删除已存在的当日数据
    today_short = trade_date_str
    existing_rows = []
    for r in range(2, ws_daily.max_row + 1):
        dval = str(ws_daily.cell(row=r, column=1).value or "")
        if dval.replace("-", "") == today_short:
            existing_rows.append(r)
    for r in reversed(existing_rows):
        ws_daily.delete_rows(r)

    # 追加新行
    next_row = ws_daily.max_row + 1
    for hd in holdings_data:
        ws_daily.cell(row=next_row, column=1).value = today_short
        ws_daily.cell(row=next_row, column=2).value = hd["name"]
        ws_daily.cell(row=next_row, column=3).value = str(hd["code"])
        ws_daily.cell(row=next_row, column=4).value = hd["quantity"] if hd["quantity"] else None
        ws_daily.cell(row=next_row, column=5).value = hd["cost_price"] if hd["cost_price"] else None
        ws_daily.cell(row=next_row, column=6).value = hd["cost_value"] if hd["cost_value"] else None
        ws_daily.cell(row=next_row, column=7).value = hd["close"]
        ws_daily.cell(row=next_row, column=8).value = round(hd["pct_chg"] / 100, 4) if hd["pct_chg"] else None
        ws_daily.cell(row=next_row, column=9).value = hd["market_value"]
        ws_daily.cell(row=next_row, column=10).value = hd["pnl"]
        ws_daily.cell(row=next_row, column=11).value = hd["pnl_pct"]
        ws_daily.cell(row=next_row, column=12).value = hd["target_weight"]
        ws_daily.cell(row=next_row, column=13).value = hd["weight"]
        ws_daily.cell(row=next_row, column=14).value = hd["deviation"]
        ws_daily.cell(row=next_row, column=15).value = "—"  # 回补数据不做信号检测
        next_row += 1

    print(f"  [OK] 日报追加 {len(holdings_data)} 条 ({trade_date_str})")
    return total_market_value - total_cost_value  # total P&L


def main():
    print("回补 2026-06-16 ~ 2026-06-17 日报数据")
    print("=" * 50)

    if not FILE_PATH.exists():
        print(f"[ERROR] 文件不存在: {FILE_PATH}")
        return 1

    ts_codes = [h["ts"] for h in HOLDINGS]
    wb = load_workbook(FILE_PATH)

    for trade_date_str in ["20260616", "20260617"]:
        print(f"\n--- {trade_date_str} ---")
        print("  获取行情...")
        prices = fetch_prices_for_date(ts_codes, trade_date_str)

        if not prices:
            print(f"  [WARN] {trade_date_str} 无行情数据，跳过")
            continue

        print(f"  获取到 {len(prices)}/{len(ts_codes)} 个品种")
        for ts_code, info in prices.items():
            print(f"    {ts_code}: close={info['close']}, chg={info['pct_chg']:+.2f}%")

        total_pnl = add_daily_snapshot(wb, prices, trade_date_str)
        print(f"  总浮动盈亏: {total_pnl:+,.2f}")

    import time
    saved = False
    for attempt in range(5):
        try:
            wb.save(FILE_PATH)
            saved = True
            break
        except PermissionError:
            print(f"  [WARN] 文件被占用，第{attempt+1}次重试...")
            time.sleep(3)
    if not saved:
        # 保存到临时文件
        tmp_path = Path(str(FILE_PATH) + ".tmp")
        wb.save(tmp_path)
        print(f"\n[WARN] 原文件被占用，已保存到临时文件: {tmp_path}")
        print(f"  请关闭 Excel 后，手动将临时文件覆盖原文件")
        return 1
    print(f"\n[OK] 文件已保存: {FILE_PATH}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
