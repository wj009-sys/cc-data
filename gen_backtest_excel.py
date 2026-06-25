#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""生成回测交易流水Excel表格"""

import json
import sys
sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 读取回测结果
with open("backtest_rules_result.json", "r", encoding="utf-8") as f:
    data = json.load(f)

trades = data["trades"]
annual = data["annual_returns"]

wb = Workbook()

# ====================================
# 样式定义
# ====================================
header_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center")
body_font = Font(name="Microsoft YaHei", size=10)
green_font = Font(name="Microsoft YaHei", size=10, color="006100")
red_font = Font(name="Microsoft YaHei", size=10, color="9C0006")
bold_font = Font(name="Microsoft YaHei", bold=True, size=10)
green_bold = Font(name="Microsoft YaHei", bold=True, size=11, color="006100")
center_align = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center")
right_align = Alignment(horizontal="right", vertical="center")
thin_border = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
buy_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
sell_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
rebal_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

# ====================================
# Sheet 1: 交易流水
# ====================================
ws1 = wb.active
ws1.title = "交易流水"
titles = ["日期", "代码", "名称", "操作", "单价", "数量", "成交金额", "原因"]
for col, t in enumerate(titles, 1):
    c = ws1.cell(row=1, column=col, value=t)
    c.font, c.fill, c.alignment, c.border = header_font, header_fill, header_align, thin_border

trades_sorted = sorted(trades, key=lambda x: (x["date"], x["code"]))
row = 2
for t in trades_sorted:
    act = t["action"]
    fill = rebal_fill if "再平衡" in act else (buy_fill if "买入" in act else sell_fill)
    vals = [t["date"], t["code"], t["name"], act, t["price"], t["qty"], t["amount"], t["reason"]]
    for col, val in enumerate(vals, 1):
        c = ws1.cell(row=row, column=col, value=val)
        c.font, c.border, c.fill = body_font, thin_border, fill
        if col in (1, 2, 4, 6):
            c.alignment = center_align
        elif col in (3, 8):
            c.alignment = left_align
        elif col in (5, 7):
            c.alignment = right_align
            c.number_format = "0.0000" if col == 5 else "#,##0.00"
        if col == 6:
            c.number_format = "#,##0"
    row += 1

for i, w in enumerate([12, 8, 14, 14, 10, 10, 14, 50], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.freeze_panes = "A2"
ws1.auto_filter.ref = f"A1:H{row-1}"

# ====================================
# Sheet 2: 年度统计
# ====================================
ws2 = wb.create_sheet("年度统计")
yheaders = ["年份", "年初市值", "年末市值", "年收益率", "最大回撤", "年化波动", "年度峰值"]
for col, t in enumerate(yheaders, 1):
    c = ws2.cell(row=1, column=col, value=t)
    c.font, c.fill, c.alignment, c.border = header_font, header_fill, header_align, thin_border

for i, ar in enumerate(annual):
    r = i + 2
    vals = [ar["year"], ar["start_value"], ar["end_value"],
            ar["return_pct"] / 100, ar["max_drawdown"] / 100,
            ar["volatility_pct"] / 100, ar["peak"]]
    for col, val in enumerate(vals, 1):
        c = ws2.cell(row=r, column=col, value=val)
        c.font, c.border, c.alignment = body_font, thin_border, center_align
        if col == 1:
            c.number_format = "0"
        elif col in (2, 3, 7):
            c.number_format = "#,##0"
        elif col in (4, 5, 6):
            c.number_format = "0.00%"
            if col == 4:
                c.font = green_font if val >= 0 else red_font

# 合计行
tr = len(annual) + 2
for col, val in enumerate([4000000, annual[-1]["end_value"]], 2):
    c = ws2.cell(row=tr, column=col, value=val)
    c.font, c.border, c.alignment = bold_font, thin_border, center_align
    c.number_format = "#,##0"

ws2.cell(row=tr, column=1, value="合计").font = bold_font
ws2.cell(row=tr, column=1).border = thin_border
ws2.cell(row=tr, column=1).alignment = center_align

cagr = ws2.cell(row=tr, column=4, value=data["summary"]["cagr_pct"] / 100)
cagr.font, cagr.border, cagr.alignment = green_bold, thin_border, center_align
cagr.number_format = "0.00%"

tot = ws2.cell(row=tr, column=5, value=data["summary"]["total_return_pct"] / 100)
tot.font, tot.border, tot.alignment = bold_font, thin_border, center_align
tot.number_format = "0.00%"

ws2.cell(row=tr, column=6, value="CAGR").font = bold_font
ws2.cell(row=tr, column=6).border = thin_border
ws2.cell(row=tr, column=6).alignment = center_align

final_gain = data["summary"]["final_value"] - 4000000
ws2.cell(row=tr, column=7, value=f"+{final_gain:,.0f} 元").font = bold_font
ws2.cell(row=tr, column=7).border = thin_border
ws2.cell(row=tr, column=7).alignment = center_align

for i, w in enumerate([8, 14, 14, 10, 10, 10, 14], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ====================================
# Sheet 3: 品种年度收益
# ====================================
ws3 = wb.create_sheet("品种年度收益")
etf_returns = {
    "组合(回测)": {2021: 42.6, 2022: 4.4, 2023: 19.3, 2024: 11.5, 2025: 23.0},
    "沪深300ETF (510300)": {2021: -6.1, 2022: -21.1, 2023: -11.4, 2024: 16.5, 2025: 21.6},
    "科创50ETF (588050)": {2021: -0.9, 2022: -29.5, 2023: -13.3, 2024: 17.7, 2025: 40.5},
    "创业板ETF (159915)": {2021: 8.1, 2022: -27.7, 2023: -19.5, 2024: 16.4, 2025: 57.8},
    "红利低波ETF (512890)": {2021: -39.8, 2022: 1.1, 2023: 11.2, 2024: 21.5, 2025: 6.8},
    "可转债ETF (511380)": {2021: 13.5, 2022: -7.3, 2023: -0.6, 2024: 5.8, 2025: 18.9},
    "10年国债ETF (511260)": {2021: 5.6, 2022: 2.6, 2023: 4.3, 2024: 8.7, 2025: -1.8},
    "5年国债ETF (511010)": {2021: 4.4, 2022: 2.1, 2023: 3.3, 2024: 6.0, 2025: -0.9},
    "短融ETF (511360)": {2021: 2.8, 2022: 2.2, 2023: 2.4, 2024: 2.2, 2025: 1.5},
    "黄金ETF (518880)": {2021: -6.6, 2022: 10.1, 2023: 16.6, 2024: 27.0, 2025: 55.6},
}

h3 = ["品种"] + [str(y) for y in range(2021, 2026)]
for col, t in enumerate(h3, 1):
    c = ws3.cell(row=1, column=col, value=t)
    c.font, c.fill, c.alignment, c.border = header_font, header_fill, header_align, thin_border

r3 = 2
for name, yd in etf_returns.items():
    c = ws3.cell(row=r3, column=1, value=name)
    c.font, c.border, c.alignment = body_font, thin_border, left_align
    if name == "组合(回测)":
        c.font = Font(name="Microsoft YaHei", bold=True, size=10)
    for y in range(2021, 2026):
        ret = yd.get(y, 0) / 100
        cell = ws3.cell(row=r3, column=y - 2020, value=ret)
        cell.font, cell.border, cell.alignment = body_font, thin_border, center_align
        cell.number_format = "0.0%"
        if name == "组合(回测)":
            cell.font = Font(name="Microsoft YaHei", bold=True, size=10,
                             color="006100" if ret >= 0 else "9C0006")
        else:
            cell.font = green_font if ret >= 0 else red_font
    r3 += 1

ws3.column_dimensions["A"].width = 24
for i in range(2, 7):
    ws3.column_dimensions[get_column_letter(i)].width = 10

# ====================================
# 保存
# ====================================
filepath = "backtest_2021_2025_trades.xlsx"
wb.save(filepath)
print(f"OK: {filepath}")
print(f"  Sheet1: 交易流水 ({len(trades)} 笔)")
print(f"  Sheet2: 年度统计")
print(f"  Sheet3: 品种年度收益")
